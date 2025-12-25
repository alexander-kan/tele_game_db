"""Excel file fixtures for testing."""

from __future__ import annotations

import tempfile
from collections.abc import Iterator
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def temp_excel() -> Iterator[Path]:
    """Create a temporary Excel file with test data.

    Yields:
        Path to temporary Excel file
    """
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create workbook with test sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create init_games sheet
    init_sheet = wb.create_sheet("init_games")
    init_sheet.append(
        [
            "Game Name",
            "Platforms",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time Beat",
            "Trailer URL",
            "My Time Beat",
            "Last Launch Date",
            "Additional Time",
        ]
    )
    init_sheet.append(
        [
            "Test Game 1",
            "Steam",
            "Пройдена",
            "January 1, 2024",
            "8.0",
            "8.5",
            "9",
            "https://metacritic.com/game1",
            "10.5",
            "https://youtube.com/trailer1",
            "12.0",
            "January 15, 2024",
            "0",
        ]
    )

    # Create new_games sheet
    new_sheet = wb.create_sheet("new_games")
    new_sheet.append(
        [
            "Game Name",
            "Platforms",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time Beat",
            "Trailer URL",
            "My Time Beat",
            "Last Launch Date",
            "Additional Time",
        ]
    )

    # Create update_games sheet
    update_sheet = wb.create_sheet("update_games")
    update_sheet.append(
        [
            "Game Name",
            "Platforms",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time Beat",
            "Trailer URL",
            "My Time Beat",
            "Last Launch Date",
            "Additional Time",
        ]
    )

    wb.save(str(excel_path))

    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def empty_excel() -> Iterator[Path]:
    """Create an empty temporary Excel file.

    Yields:
        Path to temporary Excel file
    """
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create minimal workbook
    wb = Workbook()
    wb.remove(wb.active)
    init_sheet = wb.create_sheet("init_games")
    init_sheet.append(
        [
            "Game Name",
            "Platforms",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time Beat",
            "Trailer URL",
            "My Time Beat",
            "Last Launch Date",
            "Additional Time",
        ]
    )

    wb.save(str(excel_path))

    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)
