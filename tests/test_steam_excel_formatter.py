"""Tests for SteamExcelFormatter."""

from __future__ import annotations

from unittest.mock import MagicMock, Mock

import pytest
from openpyxl import Workbook

from game_db.constants import ExcelColumn, EXCEL_DATE_NOT_SET, EXCEL_NONE_VALUE
from game_db.excel.steam_formatter import SteamExcelFormatter
from game_db.types import SteamGame


@pytest.fixture
def sample_workbook() -> Workbook:
    """Create a sample workbook with init_games sheet."""
    wb = Workbook()
    wb.create_sheet("init_games")
    # Add header row
    sheet = wb["init_games"]
    sheet.cell(row=1, column=ExcelColumn.STATUS).value = "Статус"
    sheet.cell(row=1, column=ExcelColumn.MY_TIME_BEAT).value = "Время"
    sheet.cell(row=1, column=ExcelColumn.LAST_LAUNCH_DATE).value = "Дата"
    sheet.cell(row=1, column=ExcelColumn.ADDITIONAL_TIME).value = "Доп. время"
    return wb


@pytest.fixture
def sample_game_with_playtime() -> SteamGame:
    """Create a sample SteamGame with playtime."""
    return SteamGame(
        appid=12345,
        name="Test Game",
        playtime_forever=120,  # 2 hours in minutes
        img_icon_url="icon.jpg",
        img_logo_url="logo.jpg",
        has_community_visible_stats=True,
        playtime_windows_forever=120,
        playtime_mac_forever=None,
        playtime_linux_forever=None,
        rtime_last_played=1609459200,  # 2021-01-01
    )


@pytest.fixture
def sample_game_without_playtime() -> SteamGame:
    """Create a sample SteamGame without playtime."""
    return SteamGame(
        appid=67890,
        name="Unplayed Game",
        playtime_forever=0,
        img_icon_url="icon.jpg",
        img_logo_url="logo.jpg",
        has_community_visible_stats=False,
        playtime_windows_forever=None,
        playtime_mac_forever=None,
        playtime_linux_forever=None,
        rtime_last_played=None,
    )


class TestSteamExcelFormatter:
    """Test SteamExcelFormatter methods."""

    def test_format_playtime_hours(self) -> None:
        """Test playtime formatting from minutes to hours."""
        assert SteamExcelFormatter.format_playtime_hours(60) == "1.0"
        assert SteamExcelFormatter.format_playtime_hours(90) == "1.5"
        assert SteamExcelFormatter.format_playtime_hours(120) == "2.0"
        assert SteamExcelFormatter.format_playtime_hours(75) == "1.25"

    def test_update_game_with_playtime(
        self,
        sample_workbook: Workbook,
        sample_game_with_playtime: SteamGame,
    ) -> None:
        """Test updating Excel row with game playtime."""
        sheet = sample_workbook["init_games"]
        row_number = 2

        # Set initial status
        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Не начата"

        # Mock date formatter
        date_formatter = Mock(return_value="January 1, 2021")

        # Update game
        SteamExcelFormatter.update_game_with_playtime(
            sheet, row_number, sample_game_with_playtime, date_formatter
        )

        # Verify updates
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == "2.0"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value
            == "January 1, 2021"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Брошена"
        )
        date_formatter.assert_called_once_with(1609459200)

    def test_update_game_with_playtime_no_last_played(
        self,
        sample_workbook: Workbook,
    ) -> None:
        """Test updating game with playtime but no last played date."""
        sheet = sample_workbook["init_games"]
        row_number = 2

        game = SteamGame(
            appid=12345,
            name="Test Game",
            playtime_forever=180,
            img_icon_url="icon.jpg",
            img_logo_url="logo.jpg",
            has_community_visible_stats=True,
            playtime_windows_forever=180,
            playtime_mac_forever=None,
            playtime_linux_forever=None,
            rtime_last_played=None,  # No last played date
        )

        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Не начата"
        date_formatter = Mock()

        SteamExcelFormatter.update_game_with_playtime(
            sheet, row_number, game, date_formatter
        )

        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == "3.0"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value
            is None
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Брошена"
        )
        date_formatter.assert_not_called()

    def test_reset_game_without_playtime(
        self,
        sample_workbook: Workbook,
    ) -> None:
        """Test resetting game row when playtime is zero."""
        sheet = sample_workbook["init_games"]
        row_number = 2

        # Set additional_time to sentinel none value
        sheet.cell(row=row_number, column=ExcelColumn.ADDITIONAL_TIME).value = (
            EXCEL_NONE_VALUE
        )
        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Брошена"

        SteamExcelFormatter.reset_game_without_playtime(sheet, row_number)

        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == EXCEL_NONE_VALUE
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value
            == EXCEL_DATE_NOT_SET
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Не начата"
        )

    def test_reset_game_without_playtime_skips_if_additional_time_not_none(
        self,
        sample_workbook: Workbook,
    ) -> None:
        """Test that reset is skipped if additional_time is not sentinel."""
        sheet = sample_workbook["init_games"]
        row_number = 2

        # Set additional_time to something other than sentinel
        sheet.cell(row=row_number, column=ExcelColumn.ADDITIONAL_TIME).value = "10"
        sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value = "5.0"
        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Брошена"

        SteamExcelFormatter.reset_game_without_playtime(sheet, row_number)

        # Values should remain unchanged
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == "5.0"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Брошена"
        )

    def test_update_game_row_with_playtime(
        self,
        sample_workbook: Workbook,
        sample_game_with_playtime: SteamGame,
    ) -> None:
        """Test update_game_row with game that has playtime."""
        sheet = sample_workbook["init_games"]
        row_number = 2
        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Не начата"

        date_formatter = Mock(return_value="January 1, 2021")

        SteamExcelFormatter.update_game_row(
            sheet, row_number, sample_game_with_playtime, date_formatter
        )

        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == "2.0"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value
            == "January 1, 2021"
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Брошена"
        )

    def test_update_game_row_without_playtime(
        self,
        sample_workbook: Workbook,
        sample_game_without_playtime: SteamGame,
    ) -> None:
        """Test update_game_row with game that has no playtime."""
        sheet = sample_workbook["init_games"]
        row_number = 2
        sheet.cell(row=row_number, column=ExcelColumn.ADDITIONAL_TIME).value = (
            EXCEL_NONE_VALUE
        )
        sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Брошена"

        date_formatter = Mock()

        SteamExcelFormatter.update_game_row(
            sheet, row_number, sample_game_without_playtime, date_formatter
        )

        assert (
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value
            == EXCEL_NONE_VALUE
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value
            == EXCEL_DATE_NOT_SET
        )
        assert (
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
            == "Не начата"
        )
        date_formatter.assert_not_called()
