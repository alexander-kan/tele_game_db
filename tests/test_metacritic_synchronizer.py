"""Integration tests for MetacriticSynchronizer with mocked dependencies."""

# pylint: disable=redefined-outer-name

from __future__ import annotations

import pathlib
import sys
import tempfile
from collections.abc import Iterator
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook

from game_db.config import DBFilesConfig, Paths, SettingsConfig
from game_db.db import DatabaseManager, MetacriticSynchronizer
from game_db.db_excel import ExcelImporter

# Ensure project root is on sys.path when running tests directly
PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


@pytest.fixture
def mock_excel_file_with_url() -> Iterator[Path]:
    """Create a temporary Excel file with Metacritic URL."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Test game row with Metacritic URL
    ws.append(
        [
            "Test Game",
            "Steam",
            "Не начата",
            "January 1, 2024",
            "8.0",
            "7.5",
            "8",
            "https://www.metacritic.com/game/pc/test-game",
            "10.5",
            "https://youtube.com/trailer",
            "none",
            "December 12, 4712",
            "none",
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def mock_excel_file_without_url() -> Iterator[Path]:
    """Create a temporary Excel file without Metacritic URL."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Test game row without Metacritic URL
    ws.append(
        [
            "Test Game Without URL",
            "Steam",
            "Не начата",
            "January 1, 2024",
            "8.0",
            "7.5",
            "8",
            None,  # No URL
            "10.5",
            "https://youtube.com/trailer",
            "none",
            "December 12, 4712",
            "none",
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def mock_configs() -> SettingsConfig:
    """Create mock configuration objects."""
    paths = Paths(
        backup_dir=Path("/tmp/backup"),
        update_db_dir=Path("/tmp/update_db"),
        files_dir=Path("/tmp/files"),
        sql_root=Path("/tmp/sql"),
        sqlite_db_file=Path("test.db"),
        games_excel_file=Path("/tmp/backup/games.xlsx"),
    )
    db_files = DBFilesConfig(
        sql_drop_tables=Path("test_drop.sql"),
        sql_create_tables=Path("test_create.sql"),
        sql_dictionaries=Path("test_dict.sql"),
        sql_games=Path("games.sql"),
        sql_games_on_platforms=Path("games_on_platforms.sql"),
        sqlite_db_file=Path("test.db"),
    )
    settings = SettingsConfig(paths=paths, db_files=db_files)

    return settings


@pytest.fixture
def mock_metacritic_scraper_data() -> dict:
    """Mock data returned by MetaCriticScraper."""
    return {
        "url": "https://www.metacritic.com/game/pc/test-game",
        "title": "Test Game",
        "release_date": "Aug 7, 2020",
        "critic_score": "88",
        "user_score": "8.8",
        "platform": "PC",
        "publisher": "Test Publisher",
        "developer": "Test Developer",
        "genre": "Action",
        "rating": "T",
    }


@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_with_url(
    mock_scraper_class: Mock,
    mock_excel_file_with_url: Path,
    mock_configs: SettingsConfig,
    mock_metacritic_scraper_data: dict,
) -> None:
    """Test synchronize_metacritic_games updates Excel when URL exists."""
    settings = mock_configs

    # Mock MetaCriticScraper
    mock_scraper = Mock()
    mock_scraper.game = mock_metacritic_scraper_data
    mock_scraper_class.return_value = mock_scraper

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = MetacriticSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_metacritic_games(str(mock_excel_file_with_url))

    # Verify MetaCriticScraper was called with existing URL
    mock_scraper_class.assert_called_once()
    call_url = mock_scraper_class.call_args[0][0]
    assert "metacritic.com" in call_url

    # Verify DB recreation was attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result
    assert result is True


@patch("game_db.db.search_metacritic_game_url")
@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_without_url(
    mock_scraper_class: Mock,
    mock_search: Mock,
    mock_excel_file_without_url: Path,
    mock_configs: SettingsConfig,
    mock_metacritic_scraper_data: dict,
) -> None:
    """Test synchronize_metacritic_games searches for URL when not provided."""
    settings = mock_configs

    # Mock search to return URL
    mock_search.return_value = "https://www.metacritic.com/game/pc/test-game-without-url"

    # Mock MetaCriticScraper
    mock_scraper = Mock()
    mock_scraper.game = mock_metacritic_scraper_data
    mock_scraper_class.return_value = mock_scraper

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = MetacriticSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_metacritic_games(str(mock_excel_file_without_url))

    # Verify search was called
    mock_search.assert_called_once()
    assert "Test Game Without URL" in mock_search.call_args[0][0]

    # Verify MetaCriticScraper was called with found URL
    mock_scraper_class.assert_called_once()
    call_url = mock_scraper_class.call_args[0][0]
    assert mock_search.return_value in call_url

    # Verify DB recreation was attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result
    assert result is True


@patch("game_db.db.search_metacritic_game_url")
@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_search_fails(
    mock_scraper_class: Mock,
    mock_search: Mock,
    mock_excel_file_without_url: Path,
    mock_configs: SettingsConfig,
) -> None:
    """Test handling when search fails."""
    settings = mock_configs

    # Mock search to return None (not found)
    mock_search.return_value = None

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = MetacriticSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_metacritic_games(str(mock_excel_file_without_url))

    # Verify search was called
    mock_search.assert_called_once()

    # Verify MetaCriticScraper was NOT called (no URL found)
    mock_scraper_class.assert_not_called()

    # Verify DB recreation was still attempted (even with no updates)
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result
    assert result is True


@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_test_mode_limit(
    mock_scraper_class: Mock,
    mock_excel_file_with_url: Path,
    mock_configs: SettingsConfig,
    mock_metacritic_scraper_data: dict,
) -> None:
    """Test test_mode limits to 20 games."""
    settings = mock_configs

    # Create Excel with more than 20 games
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Add 25 games
    for i in range(25):
        ws.append(
            [
                f"Game {i}",
                "Steam",
                "Не начата",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                f"https://www.metacritic.com/game/pc/game-{i}",
                "10.5",
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

    wb.save(str(mock_excel_file_with_url))

    # Mock MetaCriticScraper
    mock_scraper = Mock()
    mock_scraper.game = mock_metacritic_scraper_data
    mock_scraper_class.return_value = mock_scraper

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = MetacriticSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,  # Enable test mode
    )

    result = synchronizer.synchronize_metacritic_games(str(mock_excel_file_with_url))

    # Verify only 20 games were processed (test mode limit)
    assert mock_scraper_class.call_count == 20

    # Verify result
    assert result is True


@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_no_games_found(
    mock_scraper_class: Mock,
    mock_configs: SettingsConfig,
) -> None:
    """Test handling when no games are found in Excel."""
    settings = mock_configs

    # Create empty Excel file
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"
    # Only header, no games
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )
    wb.save(str(excel_path))

    try:
        mock_excel_importer = Mock()
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = MetacriticSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
            test_mode=True,
        )

        result = synchronizer.synchronize_metacritic_games(str(excel_path))

        # Verify MetaCriticScraper was NOT called
        mock_scraper_class.assert_not_called()

        # Verify result is False (no games found)
        assert result is False
    finally:
        excel_path.unlink(missing_ok=True)


@patch("game_db.db.MetaCriticScraper")
def test_synchronize_metacritic_games_scraper_fails(
    mock_scraper_class: Mock,
    mock_excel_file_with_url: Path,
    mock_configs: SettingsConfig,
) -> None:
    """Test handling when MetaCriticScraper fails."""
    settings = mock_configs

    # Mock MetaCriticScraper to raise exception
    mock_scraper_class.side_effect = Exception("Scraping failed")

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = MetacriticSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_metacritic_games(str(mock_excel_file_with_url))

    # Verify MetaCriticScraper was called
    mock_scraper_class.assert_called_once()

    # Verify DB recreation was still attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result (should still succeed even if scraping fails for some games)
    assert result is True


class TestMetacriticSynchronizerPrivateMethods:
    """Test private methods of MetacriticSynchronizer for better unit test coverage."""

    def test_load_workbook(
        self,
        mock_excel_file_with_url: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _load_workbook method."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = MetacriticSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_url))

        assert workbook is not None
        assert "init_games" in workbook.sheetnames

    def test_get_games_for_sync_with_url(
        self,
        mock_excel_file_with_url: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _get_games_for_sync with games that have URLs."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = MetacriticSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_url))
        games = synchronizer._get_games_for_sync(workbook)

        assert len(games) == 1
        url, _, has_url = games[0]
        assert has_url is True
        assert "metacritic.com" in url

    def test_get_games_for_sync_without_url(
        self,
        mock_excel_file_without_url: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _get_games_for_sync with games that don't have URLs."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = MetacriticSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_without_url))
        games = synchronizer._get_games_for_sync(workbook)

        assert len(games) == 1
        game_name, _, has_url = games[0]
        assert has_url is False
        assert game_name == "Test Game Without URL"
