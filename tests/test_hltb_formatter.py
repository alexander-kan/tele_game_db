"""Unit tests for HowLongToBeat Excel formatter."""

from __future__ import annotations

from openpyxl import Workbook

from game_db.constants import ExcelColumn
from game_db.excel.hltb_formatter import HowLongToBeatExcelFormatter


class TestHowLongToBeatExcelFormatter:
    """Test HowLongToBeatExcelFormatter class."""

    def test_format_time_valid(self) -> None:
        """Test formatting valid time value."""
        result = HowLongToBeatExcelFormatter.format_time(12.5)
        assert result == "12.5"

    def test_format_time_zero(self) -> None:
        """Test formatting zero time value."""
        result = HowLongToBeatExcelFormatter.format_time(0.0)
        assert result is None

    def test_format_time_negative(self) -> None:
        """Test formatting negative time value."""
        result = HowLongToBeatExcelFormatter.format_time(-5.0)
        assert result is None

    def test_format_time_none(self) -> None:
        """Test formatting None time value."""
        result = HowLongToBeatExcelFormatter.format_time(None)
        assert result is None

    def test_update_game_row_with_data(self) -> None:
        """Test updating Excel row with HowLongToBeat data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,  # Empty average_time_beat
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        hltb_data = {
            "game_name": "Test Game",
            "game_id": "12345",
            "main_story": 12.5,
            "main_extra": 18.0,
            "completionist": 25.0,
            "all_styles": 20.0,
            "similarity": 0.95,
        }

        HowLongToBeatExcelFormatter.update_game_row(ws, 2, hltb_data, partial_mode=False)

        # Verify update
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "12.5"

    def test_update_game_row_not_found_partial_mode(self) -> None:
        """Test updating Excel row when game not found in partial mode."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row with empty average_time_beat
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,  # Empty
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        # Game not found (hltb_data is None)
        HowLongToBeatExcelFormatter.update_game_row(ws, 2, None, partial_mode=True)

        # Verify "0" was written (partial mode)
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "0"

    def test_update_game_row_not_found_full_mode_empty(self) -> None:
        """Test updating Excel row when game not found in full mode with empty field."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row with empty average_time_beat
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,  # Empty
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        # Game not found (hltb_data is None)
        HowLongToBeatExcelFormatter.update_game_row(ws, 2, None, partial_mode=False)

        # Verify "0" was written (field was empty)
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "0"

    def test_update_game_row_not_found_full_mode_preserve(self) -> None:
        """Test updating Excel row when game not found in full mode preserves existing value."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row with filled average_time_beat
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                "15.5",  # Already filled
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        original_value = ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value

        # Game not found (hltb_data is None)
        HowLongToBeatExcelFormatter.update_game_row(ws, 2, None, partial_mode=False)

        # Verify original value was preserved
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == original_value
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "15.5"

    def test_update_game_row_uses_main_story(self) -> None:
        """Test that main_story is used when available."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        hltb_data = {
            "game_name": "Test Game",
            "game_id": "12345",
            "main_story": 10.0,  # Should use this
            "main_extra": 15.0,
            "completionist": 20.0,
            "all_styles": 18.0,
            "similarity": 0.95,
        }

        HowLongToBeatExcelFormatter.update_game_row(ws, 2, hltb_data, partial_mode=False)

        # Verify main_story was used
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "10.0"

    def test_update_game_row_uses_completionist_when_main_story_missing(self) -> None:
        """Test that completionist is used when main_story is missing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        hltb_data = {
            "game_name": "Test Game",
            "game_id": "12345",
            "main_story": None,  # Missing
            "main_extra": None,
            "completionist": 25.0,  # Should use this
            "all_styles": 20.0,
            "similarity": 0.95,
        }

        HowLongToBeatExcelFormatter.update_game_row(ws, 2, hltb_data, partial_mode=False)

        # Verify completionist was used
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == "25.0"

    def test_update_game_row_no_valid_time(self) -> None:
        """Test updating Excel row when no valid time data is available."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game/test",
                None,
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        original_value = ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value

        hltb_data = {
            "game_name": "Test Game",
            "game_id": "12345",
            "main_story": None,
            "main_extra": None,
            "completionist": None,
            "all_styles": None,
            "similarity": 0.95,
        }

        HowLongToBeatExcelFormatter.update_game_row(ws, 2, hltb_data, partial_mode=False)

        # Verify value remains unchanged (None)
        assert ws.cell(row=2, column=ExcelColumn.AVERAGE_TIME_BEAT).value == original_value
