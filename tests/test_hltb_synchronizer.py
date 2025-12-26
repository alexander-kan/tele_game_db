"""Integration tests for HowLongToBeatSynchronizer with mocked dependencies."""

# pylint: disable=redefined-outer-name

from __future__ import annotations

import pathlib
import sys
import tempfile
from collections.abc import Iterator
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook

from game_db.config import DBFilesConfig, Paths, SettingsConfig
from game_db.db import DatabaseManager, HowLongToBeatSynchronizer
from game_db.db_excel import ExcelImporter

# Ensure project root is on sys.path when running tests directly
PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


@pytest.fixture
def mock_excel_file_with_empty_time() -> Iterator[Path]:
    """Create a temporary Excel file with empty average_time_beat."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Test game row with empty average_time_beat
    ws.append(
        [
            "Test Game",
            "Steam",
            "Not Started",
            "January 1, 2024",
            "8.0",
            "7.5",
            "8",
            "https://www.metacritic.com/game/pc/test-game",
            None,  # Empty average_time_beat
            "https://youtube.com/trailer",
            "none",
            "December 12, 4712",
            "none",
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def mock_excel_file_with_time() -> Iterator[Path]:
    """Create a temporary Excel file with filled average_time_beat."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Test game row with filled average_time_beat
    ws.append(
        [
            "Test Game With Time",
            "Steam",
            "Not Started",
            "January 1, 2024",
            "8.0",
            "7.5",
            "8",
            "https://www.metacritic.com/game/pc/test-game",
            "15.5",  # Already filled
            "https://youtube.com/trailer",
            "none",
            "December 12, 4712",
            "none",
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def mock_configs() -> SettingsConfig:
    """Create mock configuration objects."""
    paths = Paths(
        backup_dir=Path("/tmp/backup"),
        update_db_dir=Path("/tmp/update_db"),
        files_dir=Path("/tmp/files"),
        sql_root=Path("/tmp/sql"),
        sqlite_db_file=Path("test.db"),
        games_excel_file=Path("/tmp/backup/games.xlsx"),
    )
    db_files = DBFilesConfig(
        sql_drop_tables=Path("test_drop.sql"),
        sql_create_tables=Path("test_create.sql"),
        sql_dictionaries=Path("test_dict.sql"),
        sql_games=Path("games.sql"),
        sql_games_on_platforms=Path("games_on_platforms.sql"),
        sqlite_db_file=Path("test.db"),
    )
    settings = SettingsConfig(paths=paths, db_files=db_files, owner_name="Alexander")

    return settings


@pytest.fixture
def mock_hltb_data() -> dict:
    """Mock data returned by HowLongToBeatClient."""
    return {
        "game_name": "Test Game",
        "game_id": "12345",
        "main_story": 12.5,
        "main_extra": 18.0,
        "completionist": 25.0,
        "all_styles": 20.0,
        "similarity": 0.95,
    }


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_success(
    mock_client_class: Mock,
    mock_excel_file_with_empty_time: Path,
    mock_configs: SettingsConfig,
    mock_hltb_data: dict,
) -> None:
    """Test synchronize_hltb_games updates Excel when game is found."""
    settings = mock_configs

    # Mock HowLongToBeatClient
    mock_client = Mock()
    mock_client.search_game.return_value = mock_hltb_data
    mock_client_class.return_value = mock_client

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = HowLongToBeatSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_hltb_games(str(mock_excel_file_with_empty_time))

    # Verify HowLongToBeatClient was called
    mock_client.search_game.assert_called_once()
    assert "Test Game" in mock_client.search_game.call_args[0][0]

    # Verify DB recreation was attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result
    assert result is True


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_not_found(
    mock_client_class: Mock,
    mock_excel_file_with_empty_time: Path,
    mock_configs: SettingsConfig,
) -> None:
    """Test synchronize_hltb_games handles game not found."""
    settings = mock_configs

    # Mock HowLongToBeatClient to return None (game not found)
    mock_client = Mock()
    mock_client.search_game.return_value = None
    mock_client_class.return_value = mock_client

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = HowLongToBeatSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_hltb_games(str(mock_excel_file_with_empty_time))

    # Verify HowLongToBeatClient was called
    mock_client.search_game.assert_called_once()

    # Verify DB recreation was still attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result (should still succeed even if game not found)
    assert result is True


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_partial_mode(
    mock_client_class: Mock,
    mock_excel_file_with_time: Path,
    mock_configs: SettingsConfig,
) -> None:
    """Test partial_mode only processes games with empty average_time_beat."""
    settings = mock_configs

    # Mock HowLongToBeatClient
    mock_client = Mock()
    mock_client.search_game.return_value = None
    mock_client_class.return_value = mock_client

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = HowLongToBeatSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    # In partial mode, game with filled time should be skipped
    result = synchronizer.synchronize_hltb_games(
        str(mock_excel_file_with_time), partial_mode=True
    )

    # Verify HowLongToBeatClient was NOT called (game skipped)
    mock_client.search_game.assert_not_called()

    # Verify result is None (no games to sync)
    assert result is None


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_test_mode_limit(
    mock_client_class: Mock,
    mock_configs: SettingsConfig,
    mock_hltb_data: dict,
) -> None:
    """Test test_mode limits to 20 games."""
    settings = mock_configs

    # Create Excel with more than 20 games
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Add 25 games with empty average_time_beat
    for i in range(25):
        ws.append(
            [
                f"Game {i}",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://www.metacritic.com/game/pc/game",
                None,  # Empty average_time_beat
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

    wb.save(str(excel_path))

    try:
        # Mock HowLongToBeatClient
        mock_client = Mock()
        mock_client.search_game.return_value = mock_hltb_data
        mock_client_class.return_value = mock_client

        # Create mock ExcelImporter and DatabaseManager
        mock_excel_importer = Mock()
        mock_excel_importer.add_games.return_value = True

        mock_db_manager = Mock(spec=DatabaseManager)
        mock_db_manager.execute_scripts_from_sql_file = Mock()

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
            test_mode=True,  # Enable test mode
        )

        result = synchronizer.synchronize_hltb_games(str(excel_path))

        # Verify only 20 games were processed (test mode limit)
        assert mock_client.search_game.call_count == 20

        # Verify result
        assert result is True
    finally:
        excel_path.unlink(missing_ok=True)


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_no_games_found(
    mock_client_class: Mock,
    mock_configs: SettingsConfig,
) -> None:
    """Test handling when no games are found in Excel."""
    settings = mock_configs

    # Create empty Excel file
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"
    # Only header, no games
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )
    wb.save(str(excel_path))

    try:
        mock_excel_importer = Mock()
        mock_db_manager = Mock(spec=DatabaseManager)

        # Mock client instance
        mock_client = Mock()
        mock_client_class.return_value = mock_client

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
            test_mode=True,
        )

        result = synchronizer.synchronize_hltb_games(str(excel_path))

        # Verify search_game was NOT called (no games to process)
        mock_client.search_game.assert_not_called()

        # Verify result is None (no games found)
        assert result is None
    finally:
        excel_path.unlink(missing_ok=True)


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_client_fails(
    mock_client_class: Mock,
    mock_excel_file_with_empty_time: Path,
    mock_configs: SettingsConfig,
) -> None:
    """Test handling when HowLongToBeatClient fails."""
    settings = mock_configs

    # Mock HowLongToBeatClient to raise exception
    mock_client = Mock()
    mock_client.search_game.side_effect = Exception("Client failed")
    mock_client_class.return_value = mock_client

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = HowLongToBeatSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    result = synchronizer.synchronize_hltb_games(str(mock_excel_file_with_empty_time))

    # Verify HowLongToBeatClient was called
    mock_client.search_game.assert_called_once()

    # Verify DB recreation was still attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result (should still succeed even if client fails for some games)
    assert result is True


@patch("game_db.db.HowLongToBeatClient")
def test_synchronize_hltb_games_partial_mode_empty_field(
    mock_client_class: Mock,
    mock_excel_file_with_empty_time: Path,
    mock_configs: SettingsConfig,
    mock_hltb_data: dict,
) -> None:
    """Test partial_mode processes games with empty average_time_beat."""
    settings = mock_configs

    # Mock HowLongToBeatClient
    mock_client = Mock()
    mock_client.search_game.return_value = mock_hltb_data
    mock_client_class.return_value = mock_client

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = HowLongToBeatSynchronizer(
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        test_mode=True,
    )

    # In partial mode, game with empty time should be processed
    result = synchronizer.synchronize_hltb_games(
        str(mock_excel_file_with_empty_time), partial_mode=True
    )

    # Verify HowLongToBeatClient was called
    mock_client.search_game.assert_called_once()

    # Verify result
    assert result is True


class TestHowLongToBeatSynchronizerPrivateMethods:
    """Test private methods of HowLongToBeatSynchronizer for better unit test coverage."""

    def test_load_workbook(
        self,
        mock_excel_file_with_empty_time: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _load_workbook method."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_empty_time))

        assert workbook is not None
        assert "init_games" in workbook.sheetnames

    def test_get_games_for_sync_full_mode(
        self,
        mock_excel_file_with_empty_time: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _get_games_for_sync in full mode (all games)."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_empty_time))
        games = synchronizer._get_games_for_sync(workbook, partial_mode=False)

        assert len(games) == 1
        game_name, row_number = games[0]
        assert game_name == "Test Game"
        assert row_number == 2

    def test_get_games_for_sync_partial_mode_empty(
        self,
        mock_excel_file_with_empty_time: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _get_games_for_sync in partial mode with empty field."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_empty_time))
        games = synchronizer._get_games_for_sync(workbook, partial_mode=True)

        # Should include game with empty average_time_beat
        assert len(games) == 1
        game_name, row_number = games[0]
        assert game_name == "Test Game"
        assert row_number == 2

    def test_get_games_for_sync_partial_mode_filled(
        self,
        mock_excel_file_with_time: Path,
        mock_configs: SettingsConfig,
    ) -> None:
        """Test _get_games_for_sync in partial mode with filled field."""
        settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = HowLongToBeatSynchronizer(
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file_with_time))
        games = synchronizer._get_games_for_sync(workbook, partial_mode=True)

        # Should skip game with filled average_time_beat
        assert len(games) == 0
