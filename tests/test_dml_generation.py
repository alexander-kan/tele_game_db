"""Tests for DML SQL file generation from Excel files."""

from __future__ import annotations

import configparser
import re
import tempfile
import uuid
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import Workbook

from game_db.config import DBFilesConfig, Paths, SettingsConfig
from game_db.constants import EXCEL_DATE_NOT_SET
from game_db.db import DatabaseManager
from game_db.db_excel import ExcelImporter


@pytest.fixture
def mock_excel_file() -> Path:
    """Create a temporary Excel file with test game data."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
            "Additional Time",
        ]
    )

    # Test game rows
    ws.append(
        [
            "Test Game 1",
            "Steam",
            "Completed",
            "May 2, 2024",
            "8.5",
            "8.0",
            "9",
            "https://www.metacritic.com/game/pc/test-game-1",
            "12.5",
            "https://youtube.com/trailer1",
            "15.0",
            "June 1, 2024",
            "2.0",
        ]
    )

    ws.append(
        [
            "Test Game 2",
            "Steam,Switch",
            "Not Started",
            "January 15, 2023",
            "7.0",
            "6.5",
            "none",
            "https://www.metacritic.com/game/pc/test-game-2",
            "20.0",
            "https://youtube.com/trailer2",
            "none",
            EXCEL_DATE_NOT_SET,
            "none",
        ]
    )

    ws.append(
        [
            "Test Game 3",
            "Switch",
            "Dropped",
            "December 12, 2022",
            "9.0",
            "8.5",
            "7",
            None,
            "5.5",
            None,
            "3.0",
            "March 10, 2023",
            "none",
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def excel_importer() -> ExcelImporter:
    """Create ExcelImporter instance for testing."""
    settings = SettingsConfig(
        paths=Paths(
            backup_dir=Path("/tmp/backup"),
            update_db_dir=Path("/tmp/update_db"),
            files_dir=Path("/tmp/files"),
            sql_root=Path("/tmp/sql"),
            sqlite_db_file=Path("test.db"),
            games_excel_file=Path("/tmp/backup/games.xlsx"),
        ),
        db_files=DBFilesConfig(
            sql_games=Path("games.sql"),
            sql_games_on_platforms=Path("games_on_platforms.sql"),
            sql_dictionaries=Path("dictionaries.sql"),
            sql_drop_tables=Path("drop_tables.sql"),
            sql_create_tables=Path("create_tables.sql"),
            sqlite_db_file=Path("test.db"),
        ),
        owner_name="Alexander",
    )

    # Create minimal config parsers
    table_names = configparser.ConfigParser()
    table_names.read_string(
        """
[TABLES]
games = games
games_on_platforms = games_on_platforms
status_dictionary = status_dictionary
platform_dictionary = platform_dictionary
"""
    )

    column_table_names = configparser.ConfigParser()
    column_table_names.read_string(
        """
[status_dictionary]
status_name = status_name
[platform_dictionary]
platform_name = platform_name
"""
    )

    values_dictionaries = configparser.ConfigParser()
    values_dictionaries.read_string(
        """
[STATUS]
pass = Completed
not_started = Not Started
abandoned = Dropped

[PLATFORM]
not_defined = NOT DEFINED
steam = Steam
switch = Switch
ps4 = PlayStation 4
ps_vita = PlayStation Vita
pc_origin = PC Origin
pc_gog = PC GOG
ps5 = PlayStation 5
n3ds = Nintendo 3DS
"""
    )

    db_manager = Mock(spec=DatabaseManager)

    return ExcelImporter(
        settings, table_names, column_table_names, values_dictionaries, db_manager
    )


class TestDMLGamesGeneration:
    """Test generation of dml_games.sql from Excel."""

    def test_generate_dml_games_sql_creates_file(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that generate_dml_games_sql creates SQL file."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                mock_excel_file, sql_path
            )

            # Verify file was created
            assert sql_path.exists()

            # Verify game_id_map was returned
            assert isinstance(game_id_map, dict)
            assert len(game_id_map) == 3  # 3 games
            assert "Test Game 1" in game_id_map
            assert "Test Game 2" in game_id_map
            assert "Test Game 3" in game_id_map

            # Verify all game_ids are valid UUIDs
            for game_id in game_id_map.values():
                uuid.UUID(game_id)  # Will raise if invalid
        finally:
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_format(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that generated SQL file has correct format."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            excel_importer.generate_dml_games_sql(mock_excel_file, sql_path)

            content = sql_path.read_text(encoding="utf-8")

            # Verify header
            assert "INSERT INTO games" in content
            assert "game_id, game_name, status, release_date" in content
            assert "VALUES" in content

            # Verify structure: should have 3 INSERT statements (one per game)
            # Count lines with game_id pattern
            game_id_pattern = r'\("([a-f0-9-]{36})"'
            matches = re.findall(game_id_pattern, content)
            assert len(matches) == 3

            # Verify ends with semicolon
            assert content.strip().endswith(";")

            # Verify no trailing comma before semicolon
            lines = content.strip().split("\n")
            last_line = lines[-1]
            assert last_line.endswith(";")
            assert not last_line.endswith(",;")
        finally:
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_date_conversion(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that dates are converted from Excel format to DB format."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            excel_importer.generate_dml_games_sql(mock_excel_file, sql_path)

            content = sql_path.read_text(encoding="utf-8")

            # Verify date conversions
            # "May 2, 2024" -> "2024-05-02"
            assert '"2024-05-02"' in content
            # "January 15, 2023" -> "2023-01-15"
            assert '"2023-01-15"' in content
            # "December 12, 2022" -> "2022-12-12"
            assert '"2022-12-12"' in content
            # EXCEL_DATE_NOT_SET -> "4712-12-12"
            assert '"4712-12-12"' in content
        finally:
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_status_conversion(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that statuses are converted to IDs."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            excel_importer.generate_dml_games_sql(mock_excel_file, sql_path)

            content = sql_path.read_text(encoding="utf-8")

            # Verify status IDs
            # "Completed" -> "1"
            assert '"1"' in content or '", "1",' in content
            # "Not Started" -> "2"
            assert '"2"' in content or '", "2",' in content
            # "Dropped" -> "3"
            assert '"3"' in content or '", "3",' in content
        finally:
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_value_formatting(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that values are properly formatted for SQL."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            excel_importer.generate_dml_games_sql(mock_excel_file, sql_path)

            content = sql_path.read_text(encoding="utf-8")

            # Verify string values are quoted
            assert '"Test Game 1"' in content
            assert '"Test Game 2"' in content
            assert '"Test Game 3"' in content

            # Verify "none" values are quoted
            assert '"none"' in content

            # Verify float values are not quoted
            # Should have numeric values for scores and times
            assert re.search(r",\s*8\.5,", content) or re.search(
                r",\s*8\.5\s*,", content
            )
            assert re.search(r",\s*12\.5,", content) or re.search(
                r",\s*12\.5\s*,", content
            )

            # Verify URLs are quoted
            assert '"https://www.metacritic.com/game/pc/test-game-1"' in content
        finally:
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_skips_invalid_rows(
        self, excel_importer: ExcelImporter
    ) -> None:
        """Test that invalid rows are skipped."""
        excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_path = Path(excel_file.name)
        excel_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Valid game
        ws.append(
            [
                "Valid Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game",
                "10.0",
                "https://youtube.com/trailer",
                "none",
                EXCEL_DATE_NOT_SET,
                "none",
            ]
        )

        # Invalid game (empty name)
        ws.append(
            [
                "",  # Invalid: empty name
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game",
                "10.0",
                "https://youtube.com/trailer",
                "none",
                EXCEL_DATE_NOT_SET,
                "none",
            ]
        )

        wb.save(str(excel_path))

        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(excel_path, sql_path)

            # Should only have 1 game (invalid row skipped)
            assert len(game_id_map) == 1
            assert "Valid Game" in game_id_map

            content = sql_path.read_text(encoding="utf-8")
            # Should only have 1 INSERT statement
            game_id_pattern = r'\("([a-f0-9-]{36})"'
            matches = re.findall(game_id_pattern, content)
            assert len(matches) == 1
        finally:
            excel_path.unlink(missing_ok=True)
            sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_sql_overwrites_existing_file(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that existing SQL file is overwritten."""
        sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        sql_path = Path(sql_file.name)
        sql_file.close()

        # Create existing file with different content
        sql_path.write_text("OLD CONTENT", encoding="utf-8")

        try:
            excel_importer.generate_dml_games_sql(mock_excel_file, sql_path)

            content = sql_path.read_text(encoding="utf-8")

            # Verify old content is gone
            assert "OLD CONTENT" not in content

            # Verify new content is present
            assert "INSERT INTO games" in content
        finally:
            sql_path.unlink(missing_ok=True)


class TestDMLGamesOnPlatformsGeneration:
    """Test generation of dml_games_on_platforms.sql from Excel."""

    def test_generate_dml_games_on_platforms_sql_creates_file(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that generate_dml_games_on_platforms_sql creates SQL file."""
        # First generate games SQL to get game_id_map
        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                mock_excel_file, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                mock_excel_file, platforms_sql_path, game_id_map
            )

            # Verify file was created
            assert platforms_sql_path.exists()

            content = platforms_sql_path.read_text(encoding="utf-8")

            # Verify header
            assert "INSERT INTO games_on_platforms" in content
            assert "platform_id, reference_game_id" in content
            assert "VALUES" in content
        finally:
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_on_platforms_sql_platform_ids(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that platform names are converted to IDs."""
        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                mock_excel_file, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                mock_excel_file, platforms_sql_path, game_id_map
            )

            content = platforms_sql_path.read_text(encoding="utf-8")

            # Verify platform IDs
            # Steam -> 2
            assert '"2"' in content
            # Switch -> 3
            assert '"3"' in content

            # Verify game_ids are referenced
            for game_id in game_id_map.values():
                assert game_id in content
        finally:
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_on_platforms_sql_multiple_platforms(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that games with multiple platforms create multiple entries."""
        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                mock_excel_file, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                mock_excel_file, platforms_sql_path, game_id_map
            )

            content = platforms_sql_path.read_text(encoding="utf-8")

            # "Test Game 2" has "Steam,Switch" -> should create 2 entries
            test_game_2_id = game_id_map["Test Game 2"]
            # Count occurrences of this game_id
            count = content.count(test_game_2_id)
            assert count == 2  # One for Steam, one for Switch

            # Verify format: should have entries like ("2", "game_id") and ("3", "game_id")
            pattern = rf'\("2", "{test_game_2_id}"\)'
            assert re.search(pattern, content) is not None
            pattern = rf'\("3", "{test_game_2_id}"\)'
            assert re.search(pattern, content) is not None
        finally:
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_on_platforms_sql_skips_not_defined(
        self, excel_importer: ExcelImporter
    ) -> None:
        """Test that 'not_defined' platform is skipped."""
        excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_path = Path(excel_file.name)
        excel_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game with "not_defined" platform (should be skipped)
        ws.append(
            [
                "Test Game",
                "NOT DEFINED",  # This maps to platform_id=1, which is skipped
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game",
                "10.0",
                "https://youtube.com/trailer",
                "none",
                EXCEL_DATE_NOT_SET,
                "none",
            ]
        )

        wb.save(str(excel_path))

        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                excel_path, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                excel_path, platforms_sql_path, game_id_map
            )

            content = platforms_sql_path.read_text(encoding="utf-8")

            # Should have no platform entries (not_defined is skipped)
            # Only header and VALUES, but no actual entries
            lines = [line.strip() for line in content.split("\n") if line.strip()]
            # Should only have INSERT statement and VALUES, no data rows
            assert len([line for line in lines if line.startswith('("')]) == 0
        finally:
            excel_path.unlink(missing_ok=True)
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_on_platforms_sql_skips_invalid_rows(
        self, excel_importer: ExcelImporter
    ) -> None:
        """Test that invalid game rows are skipped."""
        excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        excel_path = Path(excel_file.name)
        excel_file.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Valid game
        ws.append(
            [
                "Valid Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game",
                "10.0",
                "https://youtube.com/trailer",
                "none",
                EXCEL_DATE_NOT_SET,
                "none",
            ]
        )

        # Invalid game (empty name - won't be in game_id_map)
        ws.append(
            [
                "",  # Invalid
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                "https://metacritic.com/game",
                "10.0",
                "https://youtube.com/trailer",
                "none",
                EXCEL_DATE_NOT_SET,
                "none",
            ]
        )

        wb.save(str(excel_path))

        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                excel_path, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                excel_path, platforms_sql_path, game_id_map
            )

            content = platforms_sql_path.read_text(encoding="utf-8")

            # Should only have 1 platform entry (for valid game)
            valid_game_id = game_id_map["Valid Game"]
            # Count entries with this game_id
            count = content.count(valid_game_id)
            assert count == 1  # One entry for Steam platform
        finally:
            excel_path.unlink(missing_ok=True)
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)

    def test_generate_dml_games_on_platforms_sql_format(
        self, mock_excel_file: Path, excel_importer: ExcelImporter
    ) -> None:
        """Test that generated SQL file has correct format."""
        games_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        games_sql_path = Path(games_sql_file.name)
        games_sql_file.close()

        platforms_sql_file = tempfile.NamedTemporaryFile(delete=False, suffix=".sql")
        platforms_sql_path = Path(platforms_sql_file.name)
        platforms_sql_file.close()

        try:
            game_id_map = excel_importer.generate_dml_games_sql(
                mock_excel_file, games_sql_path
            )

            excel_importer.generate_dml_games_on_platforms_sql(
                mock_excel_file, platforms_sql_path, game_id_map
            )

            content = platforms_sql_path.read_text(encoding="utf-8")

            # Verify ends with semicolon
            assert content.strip().endswith(";")

            # Verify no trailing comma before semicolon
            lines = content.strip().split("\n")
            last_line = lines[-1]
            assert last_line.endswith(";")
            assert not last_line.endswith(",;")

            # Verify format: ("platform_id", "game_id")
            pattern = r'\("(\d+)", "([a-f0-9-]{36})"\)'
            matches = re.findall(pattern, content)
            assert len(matches) > 0
        finally:
            games_sql_path.unlink(missing_ok=True)
            platforms_sql_path.unlink(missing_ok=True)
