"""Tests for ExcelReader and ExcelWriter helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from game_db.constants import ExcelColumn
from game_db.excel.models import GameRow
from game_db.excel.reader import ExcelReader
from game_db.excel.writer import ExcelWriter


def _make_workbook_with_init_sheet() -> Workbook:
    """Create an in‑memory workbook with init_games sheet."""
    wb = Workbook()
    # Use a predictable sheet name instead of the default "Sheet"
    ws = wb.active
    ws.title = "init_games"
    return wb


def test_write_game_row_from_gamerow(tmp_path: Path) -> None:
    """ExcelWriter.write_game_row writes values from GameRow correctly."""
    wb = _make_workbook_with_init_sheet()
    sheet = wb["init_games"]
    xlsx_path = tmp_path / "games.xlsx"
    wb.save(xlsx_path)

    game_row = GameRow(
        game_name="Test Game",
        platforms="Steam,Switch",
        status="Completed",
        release_date="January 1, 2024",
        press_score="8.0",
        user_score="8.5",
        my_score="9.0",
        metacritic_url="https://metacritic.com/game",
        average_time_beat="10.5",
        trailer_url="https://youtube.com/trailer",
        my_time_beat="12.0",
        last_launch_date="January 15, 2024",
        additional_time="1.5",
    )

    ExcelWriter.write_game_row(sheet, 1, game_row, xlsx_path, wb)

    assert sheet.cell(row=1, column=ExcelColumn.GAME_NAME).value == "Test Game"
    assert (
        sheet.cell(row=1, column=ExcelColumn.PLATFORMS).value
        == "Steam,Switch"
    )
    assert sheet.cell(row=1, column=ExcelColumn.STATUS).value == "Completed"
    assert (
        sheet.cell(row=1, column=ExcelColumn.RELEASE_DATE).value
        == "January 1, 2024"
    )
    assert sheet.cell(row=1, column=ExcelColumn.PRESS_SCORE).value == "8.0"
    assert sheet.cell(row=1, column=ExcelColumn.USER_SCORE).value == "8.5"
    assert sheet.cell(row=1, column=ExcelColumn.MY_SCORE).value == "9.0"
    assert sheet.cell(row=1, column=ExcelColumn.METACRITIC_URL).value == (
        "https://metacritic.com/game"
    )
    assert (
        sheet.cell(row=1, column=ExcelColumn.AVERAGE_TIME_BEAT).value
        == "10.5"
    )
    assert (
        sheet.cell(row=1, column=ExcelColumn.TRAILER_URL).value
        == "https://youtube.com/trailer"
    )
    assert sheet.cell(row=1, column=ExcelColumn.MY_TIME_BEAT).value == "12.0"
    assert (
        sheet.cell(row=1, column=ExcelColumn.LAST_LAUNCH_DATE).value
        == "January 15, 2024"
    )
    assert (
        sheet.cell(row=1, column=ExcelColumn.ADDITIONAL_TIME).value == "1.5"
    )


def test_write_game_row_from_list(tmp_path: Path) -> None:
    """ExcelWriter.write_game_row works when given a raw list."""
    wb = _make_workbook_with_init_sheet()
    sheet = wb["init_games"]
    xlsx_path = tmp_path / "games.xlsx"
    wb.save(xlsx_path)

    row_data = [
        "List Game",
        "Steam",
        "Not Started",
        "February 2, 2024",
        "7.0",
        "7.5",
        "8.0",
        "https://example.com",
        "5.0",
        "https://example.com/trailer",
        "0.0",
        "February 3, 2024",
        "0",
    ]

    ExcelWriter.write_game_row(sheet, 2, row_data, xlsx_path, wb)

    assert sheet.cell(row=2, column=ExcelColumn.GAME_NAME).value == "List Game"
    assert sheet.cell(row=2, column=ExcelColumn.PLATFORMS).value == "Steam"
    assert sheet.cell(row=2, column=ExcelColumn.STATUS).value == "Not Started"


def test_append_to_init_games_sheet_appends_at_end(tmp_path: Path) -> None:
    """append_to_init_games_sheet writes data after existing rows."""
    wb = _make_workbook_with_init_sheet()
    init_sheet = wb["init_games"]
    init_sheet.append(["Header"])
    xlsx_path = tmp_path / "games.xlsx"
    wb.save(xlsx_path)

    ExcelWriter.append_to_init_games_sheet(
        wb, ["Appended Game"], xlsx_path=xlsx_path
    )

    # New row should be at max_row and contain the provided value
    last_row = init_sheet.max_row
    assert (
        init_sheet.cell(row=last_row, column=ExcelColumn.GAME_NAME).value
        == "Appended Game"
    )


def test_save_workbook_writes_file(tmp_path: Path) -> None:
    """save_workbook persists workbook to given path."""
    wb = _make_workbook_with_init_sheet()
    file_path = tmp_path / "saved.xlsx"

    ExcelWriter.save_workbook(wb, file_path)

    assert file_path.exists()


def test_excel_reader_load_and_get_sheet(tmp_path: Path) -> None:
    """ExcelReader.load_workbook and get_sheet work together."""
    wb = _make_workbook_with_init_sheet()
    sheet = wb["init_games"]
    sheet.append(["Game Name"])
    xlsx_path = tmp_path / "games.xlsx"
    wb.save(xlsx_path)

    reader = ExcelReader()
    loaded_wb = reader.load_workbook(xlsx_path)
    loaded_sheet = reader.get_sheet(loaded_wb, "init_games")

    assert loaded_sheet.title == "init_games"
    assert loaded_sheet.max_row == sheet.max_row




def test_read_game_rows_and_find_row_by_game_name(tmp_path: Path) -> None:
    """read_game_rows returns GameRow objects and supports row lookup."""
    wb = _make_workbook_with_init_sheet()
    sheet = wb["init_games"]
    # Add header row (row 1)
    sheet.append(
        [
            "Name of the game",
            "Platform",
            "Status",
            "Release year",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic",
            "Play time (HLTB)",
            "Trailer",
            "My play time",
            "Last launch",
            "My play time (on the console when the game is also on Steam)",
        ]
    )
    # Two data rows (rows 2 and 3)
    sheet.append(
        [
            "Game 1",
            "Steam",
            "Completed",
            "January 1, 2024",
            "8",
            "8.5",
            "9",
            "https://example.com/1",
            "10.5",
            "https://example.com/t1",
            "12.0",
            "January 2, 2024",
            "0",
        ]
    )
    sheet.append(
        [
            "Game 2",
            "Switch",
            "Not Started",
            "February 1, 2024",
            "7",
            "7.5",
            "8",
            "https://example.com/2",
            "5.0",
            "https://example.com/t2",
            "0.0",
            "February 3, 2024",
            "1.5",
        ]
    )
    xlsx_path = tmp_path / "games.xlsx"
    wb.save(xlsx_path)

    reader = ExcelReader()
    loaded_wb = reader.load_workbook(xlsx_path)
    loaded_sheet = reader.get_sheet(loaded_wb, "init_games")

    game_rows = reader.read_game_rows(loaded_sheet)
    assert len(game_rows) == 2
    assert isinstance(game_rows[0], GameRow)
    assert game_rows[0].game_name == "Game 1"
    assert game_rows[1].platforms == "Switch"

    # find_row_by_game_name should locate correct row (1‑based)
    # Game 2 is in row 3 (row 1 is header, row 2 is Game 1, row 3 is Game 2)
    row_index = reader.find_row_by_game_name(loaded_sheet, "Game 2")
    assert row_index == 3
    assert (
        reader.find_row_by_game_name(loaded_sheet, "Nonexistent Game") is None
    )

