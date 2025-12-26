"""Integration tests for SteamSynchronizer with mocked dependencies."""

# pylint: disable=redefined-outer-name

from __future__ import annotations

import configparser
import pathlib
import sys
import tempfile
from collections.abc import Iterator
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import Workbook

from game_db.config import DBFilesConfig, Paths, SettingsConfig, TokensConfig
from game_db.constants import EXCEL_DATE_NOT_SET, EXCEL_NONE_VALUE
from game_db.db import DatabaseManager, SteamSynchronizer
from game_db.db_excel import ExcelImporter
from game_db.types import SteamGame

# Ensure project root is on sys.path when running tests directly
PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


@pytest.fixture
def mock_excel_file() -> Iterator[Path]:
    """Create a temporary Excel file for testing."""
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_path = Path(excel_file.name)
    excel_file.close()

    # Create Excel workbook with test data
    wb = Workbook()
    ws = wb.active
    ws.title = "init_games"

    # Header row
    ws.append(
        [
            "Game Name",
            "Platform",
            "Status",
            "Release Date",
            "Press Score",
            "User Score",
            "My Score",
            "Metacritic URL",
            "Average Time",
            "Trailer URL",
            "My Time",
            "Last Launch",
        ]
    )

    # Test game row
    ws.append(
        [
            "Test Game",
            "Steam",
            "Not Started",
            "January 1, 2024",
            "8",
            "7.5",
            "8",
            "https://metacritic.com/game/test",
            "10.5",
            "https://youtube.com/trailer",
            EXCEL_NONE_VALUE,
            EXCEL_DATE_NOT_SET,
        ]
    )

    wb.save(str(excel_path))
    yield excel_path

    # Cleanup
    excel_path.unlink(missing_ok=True)


@pytest.fixture
def mock_steam_api() -> Mock:
    """Create a mock SteamAPI client."""
    mock_client = Mock()
    mock_games = [
        SteamGame(
            appid=12345,
            name="Test Game",
            playtime_forever=120,  # 2 hours in minutes
            img_icon_url="",
            img_logo_url="",
            has_community_visible_stats=None,
            playtime_windows_forever=None,
            playtime_mac_forever=None,
            playtime_linux_forever=None,
            rtime_last_played=1704067200,  # Jan 1, 2024
        )
    ]
    mock_client.get_all_games.return_value = mock_games
    return mock_client


@pytest.fixture
def mock_configs() -> tuple[TokensConfig, SettingsConfig]:
    """Create mock configuration objects."""
    tokens = TokensConfig(
        telegram_token="tg",
        steam_key="test_key",
        steam_id="123456789",
    )

    paths = Paths(
        backup_dir=Path("/tmp/backup"),
        update_db_dir=Path("/tmp/update_db"),
        files_dir=Path("/tmp/files"),
        sql_root=Path("/tmp/sql"),
        sqlite_db_file=Path("test.db"),
        games_excel_file=Path("/tmp/backup/games.xlsx"),
    )
    db_files = DBFilesConfig(
        sql_drop_tables=Path("test_drop.sql"),
        sql_create_tables=Path("test_create.sql"),
        sql_dictionaries=Path("test_dict.sql"),
        sql_games=Path("games.sql"),
        sql_games_on_platforms=Path("games_on_platforms.sql"),
        sqlite_db_file=Path("test.db"),
    )
    settings = SettingsConfig(paths=paths, db_files=db_files, owner_name="Alexander")

    return tokens, settings


def test_synchronize_steam_games_updates_excel(
    mock_excel_file: Path,
    mock_steam_api: Mock,
    mock_configs: tuple,
) -> None:
    """Test that synchronize_steam_games updates Excel with Steam data."""
    tokens, settings = mock_configs

    # Create mock ExcelImporter and DatabaseManager
    mock_excel_importer = Mock()
    # ExcelImporter now uses reader.find_row_by_game_name
    mock_excel_importer.reader = Mock()
    mock_excel_importer.reader.find_row_by_game_name.return_value = 2
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = SteamSynchronizer(
        tokens=tokens,
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        steam_client=mock_steam_api,
    )

    result = synchronizer.synchronize_steam_games(str(mock_excel_file))

    # Verify Steam API was called
    mock_steam_api.get_all_games.assert_called_once_with("123456789")

    # Verify Excel was updated
    mock_excel_importer.reader.find_row_by_game_name.assert_called()

    # Verify DB recreation was attempted
    assert mock_db_manager.execute_scripts_from_sql_file.call_count == 3

    # Verify result (returns tuple: (success: bool, similarity_matches: list))
    assert isinstance(result, tuple)
    assert result[0] is True
    assert isinstance(result[1], list)


def test_synchronize_steam_games_game_not_found(
    mock_excel_file: Path,
    mock_steam_api: Mock,
    mock_configs: tuple,
) -> None:
    """Test handling when game from Steam is not found in Excel."""
    tokens, settings = mock_configs

    # Create game that doesn't exist in Excel
    mock_steam_api.get_all_games.return_value = [
        SteamGame(
            appid=99999,
            name="Non-Existent Game",
            playtime_forever=60,
            img_icon_url="",
            img_logo_url="",
            has_community_visible_stats=None,
            playtime_windows_forever=None,
            playtime_mac_forever=None,
            playtime_linux_forever=None,
            rtime_last_played=1704067200,
        )
    ]

    mock_excel_importer = Mock()
    # ExcelImporter now uses reader.find_row_by_game_name
    mock_excel_importer.reader = Mock()
    mock_excel_importer.reader.find_row_by_game_name.return_value = None
    mock_excel_importer.add_games.return_value = True

    mock_db_manager = Mock(spec=DatabaseManager)
    mock_db_manager.execute_scripts_from_sql_file = Mock()

    synchronizer = SteamSynchronizer(
        tokens=tokens,
        excel_importer=mock_excel_importer,
        db_manager=mock_db_manager,
        settings=settings,
        steam_client=mock_steam_api,
    )

    result = synchronizer.synchronize_steam_games(str(mock_excel_file))

    # Should still complete successfully (returns tuple: (success: bool, similarity_matches: list))
    assert isinstance(result, tuple)
    assert result[0] is True
    assert isinstance(result[1], list)
    # Verify that find_row_by_game_name was called
    assert mock_excel_importer.reader.find_row_by_game_name.called


class TestSteamSynchronizerPrivateMethods:
    """Test private methods of SteamSynchronizer for better unit test coverage."""

    def test_load_workbook(
        self,
        mock_excel_file: Path,
        mock_configs: tuple,
    ) -> None:
        """Test _load_workbook method."""
        tokens, settings = mock_configs

        mock_excel_importer = Mock(spec=ExcelImporter)
        mock_db_manager = Mock(spec=DatabaseManager)

        synchronizer = SteamSynchronizer(
            tokens=tokens,
            excel_importer=mock_excel_importer,
            db_manager=mock_db_manager,
            settings=settings,
        )

        workbook = synchronizer._load_workbook(str(mock_excel_file))

        assert workbook is not None
        assert "init_games" in workbook.sheetnames

    # ... rest of tests unchanged ...
