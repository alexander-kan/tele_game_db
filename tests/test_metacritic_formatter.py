"""Unit tests for Metacritic Excel formatter."""

from __future__ import annotations

from openpyxl import Workbook

from game_db.constants import ExcelColumn
from game_db.excel.metacritic_formatter import MetacriticExcelFormatter


class TestMetacriticExcelFormatter:
    """Test MetacriticExcelFormatter class."""

    def test_parse_release_date_full_month(self) -> None:
        """Test parsing release date with full month name."""
        result = MetacriticExcelFormatter.parse_release_date("August 7, 2020")
        assert result == "August 7, 2020"

    def test_parse_release_date_abbreviated_month(self) -> None:
        """Test parsing release date with abbreviated month."""
        result = MetacriticExcelFormatter.parse_release_date("Aug 7, 2020")
        assert result == "August 7, 2020"

    def test_parse_release_date_all_months(self) -> None:
        """Test parsing release date for all month abbreviations."""
        months = [
            ("Jan", "January"),
            ("Feb", "February"),
            ("Mar", "March"),
            ("Apr", "April"),
            ("May", "May"),
            ("Jun", "June"),
            ("Jul", "July"),
            ("Aug", "August"),
            ("Sep", "September"),
            ("Oct", "October"),
            ("Nov", "November"),
            ("Dec", "December"),
        ]

        for abbr, full in months:
            date_str = f"{abbr} 15, 2024"
            expected = f"{full} 15, 2024"
            result = MetacriticExcelFormatter.parse_release_date(date_str)
            assert result == expected

    def test_parse_release_date_empty(self) -> None:
        """Test parsing empty release date."""
        result = MetacriticExcelFormatter.parse_release_date("")
        assert result == ""

    def test_format_press_score_90(self) -> None:
        """Test formatting press score 90 to 9.0."""
        result = MetacriticExcelFormatter.format_press_score(90)
        assert result == "9.0"

    def test_format_press_score_76(self) -> None:
        """Test formatting press score 76 to 7.6."""
        result = MetacriticExcelFormatter.format_press_score(76)
        assert result == "7.6"

    def test_format_press_score_0(self) -> None:
        """Test formatting press score 0 to 0.0."""
        result = MetacriticExcelFormatter.format_press_score(0)
        assert result == "0.0"

    def test_format_press_score_100(self) -> None:
        """Test formatting press score 100 to 10.0."""
        result = MetacriticExcelFormatter.format_press_score(100)
        assert result == "10.0"

    def test_format_press_score_string(self) -> None:
        """Test formatting press score from string."""
        result = MetacriticExcelFormatter.format_press_score("85")
        assert result == "8.5"

    def test_format_press_score_none(self) -> None:
        """Test formatting None press score."""
        result = MetacriticExcelFormatter.format_press_score(None)
        assert result is None

    def test_format_press_score_empty(self) -> None:
        """Test formatting empty press score."""
        result = MetacriticExcelFormatter.format_press_score("")
        assert result is None

    def test_format_score_valid(self) -> None:
        """Test formatting valid score."""
        result = MetacriticExcelFormatter.format_score("8.5")
        assert result == "8.5"

    def test_format_score_tbd(self) -> None:
        """Test formatting 'tbd' score."""
        result = MetacriticExcelFormatter.format_score("tbd")
        assert result is None

    def test_format_score_none(self) -> None:
        """Test formatting None score."""
        result = MetacriticExcelFormatter.format_score(None)
        assert result is None

    def test_update_game_row_all_fields(self) -> None:
        """Test updating Excel row with all Metacritic data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                None,
                "10.5",
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        metacritic_data = {
            "url": "https://www.metacritic.com/game/pc/test-game",
            "release_date": "Aug 7, 2020",
            "critic_score": "88",
            "user_score": "8.8",
        }

        MetacriticExcelFormatter.update_game_row(ws, 2, metacritic_data)

        # Verify updates
        assert ws.cell(row=2, column=ExcelColumn.RELEASE_DATE).value == "August 7, 2020"
        assert ws.cell(row=2, column=ExcelColumn.PRESS_SCORE).value == "8.8"
        assert ws.cell(row=2, column=ExcelColumn.USER_SCORE).value == "8.8"
        assert (
            ws.cell(row=2, column=ExcelColumn.METACRITIC_URL).value
            == "https://www.metacritic.com/game/pc/test-game"
        )

    def test_update_game_row_partial_data(self) -> None:
        """Test updating Excel row with partial Metacritic data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "init_games"

        # Header
        ws.append(
            [
                "Game Name",
                "Platform",
                "Status",
                "Release Date",
                "Press Score",
                "User Score",
                "My Score",
                "Metacritic URL",
                "Average Time",
                "Trailer URL",
                "My Time",
                "Last Launch",
                "Additional Time",
            ]
        )

        # Game row
        ws.append(
            [
                "Test Game",
                "Steam",
                "Not Started",
                "January 1, 2024",
                "8.0",
                "7.5",
                "8",
                None,
                "10.5",
                "https://youtube.com/trailer",
                "none",
                "December 12, 4712",
                "none",
            ]
        )

        # Only URL, no scores
        metacritic_data = {
            "url": "https://www.metacritic.com/game/pc/test-game",
        }

        MetacriticExcelFormatter.update_game_row(ws, 2, metacritic_data)

        # Verify only URL was updated
        url_cell = ws.cell(row=2, column=ExcelColumn.METACRITIC_URL)
        assert url_cell.value == "https://www.metacritic.com/game/pc/test-game"
        # Other fields should remain unchanged
        date_cell = ws.cell(row=2, column=ExcelColumn.RELEASE_DATE)
        assert date_cell.value == "January 1, 2024"
