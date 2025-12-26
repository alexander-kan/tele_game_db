"""Excel formatting for Metacritic synchronization.

This module provides functionality to update Excel cells with
Metacritic game data, including release date, press score,
user score, and Metacritic URL.
"""

from __future__ import annotations

import logging

from openpyxl.worksheet.worksheet import Worksheet

from ..constants import ExcelColumn

logger = logging.getLogger("game_db.excel.metacritic")


class MetacriticExcelFormatter:
    """Format Excel cells with Metacritic game data."""

    @staticmethod
    def parse_release_date(release_date_str: str) -> str:
        """Parse release date from Metacritic format to Excel format.

        Supports two input formats:
        1. Metacritic HTML format: "Aug 7, 2020" -> "August 7, 2020"
        2. JSON-LD format: "2008-04-14" -> "April 14, 2008"

        Args:
            release_date_str: Release date string from Metacritic

        Returns:
            Formatted release date string for Excel (format: "Month DD, YYYY")
        """
        if not release_date_str:
            return ""

        # Month abbreviation to full name mapping (for HTML format)
        month_abbr_map = {
            "Jan": "January",
            "Feb": "February",
            "Mar": "March",
            "Apr": "April",
            "May": "May",
            "Jun": "June",
            "Jul": "July",
            "Aug": "August",
            "Sep": "September",
            "Oct": "October",
            "Nov": "November",
            "Dec": "December",
        }

        # Month number to full name mapping (for JSON-LD format: YYYY-MM-DD)
        month_num_map = {
            "01": "January",
            "02": "February",
            "03": "March",
            "04": "April",
            "05": "May",
            "06": "June",
            "07": "July",
            "08": "August",
            "09": "September",
            "10": "October",
            "11": "November",
            "12": "December",
        }

        # Check if it's JSON-LD format (YYYY-MM-DD)
        if "-" in release_date_str and len(release_date_str.split("-")) == 3:
            date_parts = release_date_str.split("-")
            if len(date_parts) == 3:
                year = date_parts[0]
                month_num = date_parts[1]
                day = date_parts[2]

                month_full = month_num_map.get(month_num, month_num)
                # Remove leading zero from day if present
                day = str(int(day)) if day.isdigit() else day
                return f"{month_full} {day}, {year}"

        # Otherwise, treat as HTML format (e.g., "Aug 7, 2020")
        parts = release_date_str.split()
        if len(parts) >= 3:
            month_abbr = parts[0]
            day = parts[1].rstrip(",")
            year = parts[2]

            month_full = month_abbr_map.get(month_abbr, month_abbr)
            return f"{month_full} {day}, {year}"

        return release_date_str

    @staticmethod
    def format_score(score: str | int | None) -> str | None:
        """Format score value for Excel.

        Args:
            score: Score value from Metacritic (can be string or int)

        Returns:
            Formatted score string or None
        """
        if score is None or score == "":
            return None

        # Convert to string and strip
        score_str = str(score).strip()
        if not score_str or score_str == "tbd":
            return None

        return score_str

    @staticmethod
    def format_press_score(press_score: str | int | None) -> str | None:
        """Format press score from 0-100 scale to 0-10 scale.

        Converts Metacritic press score (0-100) to match user score format (0-10).
        Examples: 90 -> 9.0, 76 -> 7.6, 0 -> 0.0, 100 -> 10.0

        Args:
            press_score: Press score value from Metacritic (0-100 scale)

        Returns:
            Formatted score string (0-10 scale) or None
        """
        if press_score is None or press_score == "":
            return None

        try:
            # Convert to float
            score_float = float(str(press_score).strip())
            if score_float == 0:
                return "0.0"
            if score_float == 100:
                return "10.0"

            # Convert from 0-100 to 0-10 scale
            converted_score = score_float / 10.0
            return str(converted_score)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def update_game_row(
        sheet: Worksheet,
        row_number: int,
        metacritic_data: dict,
    ) -> None:
        """Update Excel row with Metacritic game data.

        Updates RELEASE_DATE, PRESS_SCORE, USER_SCORE, and METACRITIC_URL columns.

        Args:
            sheet: Excel worksheet to update
            row_number: Row number (1-based) to update
            metacritic_data: Dictionary with Metacritic data from MetaCriticScraper
        """
        # Update release date
        release_date = metacritic_data.get("release_date")
        logger.info(
            "[METACRITIC_FORMATTER] Row %d: release_date raw=%r",
            row_number,
            release_date,
        )
        if release_date:
            formatted_date = MetacriticExcelFormatter.parse_release_date(
                release_date
            )
            logger.info(
                "[METACRITIC_FORMATTER] Row %d: release_date formatted=%r",
                row_number,
                formatted_date,
            )
            if formatted_date:
                sheet.cell(
                    row=row_number, column=ExcelColumn.RELEASE_DATE
                ).value = formatted_date
                logger.info(
                    "[METACRITIC_FORMATTER] Updated release_date for row %d: %s",
                    row_number,
                    formatted_date,
                )

        # Update press score (critic_score) - convert from 0-100 to 0-10 scale
        critic_score_raw = metacritic_data.get("critic_score")
        logger.info(
            "[METACRITIC_FORMATTER] Row %d: critic_score raw=%r",
            row_number,
            critic_score_raw,
        )
        press_score = MetacriticExcelFormatter.format_press_score(critic_score_raw)
        logger.info(
            "[METACRITIC_FORMATTER] Row %d: press_score formatted=%r",
            row_number,
            press_score,
        )
        if press_score is not None:
            sheet.cell(
                row=row_number, column=ExcelColumn.PRESS_SCORE
            ).value = press_score
            logger.info(
                "[METACRITIC_FORMATTER] Updated press_score for row %d: %s",
                row_number,
                press_score,
            )

        # Update user score
        user_score_raw = metacritic_data.get("user_score")
        logger.info(
            "[METACRITIC_FORMATTER] Row %d: user_score raw=%r",
            row_number,
            user_score_raw,
        )
        user_score = MetacriticExcelFormatter.format_score(user_score_raw)
        logger.info(
            "[METACRITIC_FORMATTER] Row %d: user_score formatted=%r",
            row_number,
            user_score,
        )
        if user_score is not None:
            sheet.cell(
                row=row_number, column=ExcelColumn.USER_SCORE
            ).value = user_score
            logger.info(
                "[METACRITIC_FORMATTER] Updated user_score for row %d: %s",
                row_number,
                user_score,
            )

        # Update Metacritic URL
        if "url" in metacritic_data and metacritic_data.get("url"):
            sheet.cell(
                row=row_number, column=ExcelColumn.METACRITIC_URL
            ).value = metacritic_data["url"]
