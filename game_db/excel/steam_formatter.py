"""Excel formatting for Steam synchronization.

This module provides functionality to update Excel cells with Steam game data,
including playtime, last launch dates, and status updates.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..constants import EXCEL_DATE_NOT_SET, EXCEL_NONE_VALUE, ExcelColumn
from ..types import SteamGame

if TYPE_CHECKING:
    from pathlib import Path

logger = logging.getLogger("game_db.excel.steam")


class SteamExcelFormatter:
    """Format Excel cells with Steam game data."""

    @staticmethod
    def format_playtime_hours(playtime_minutes: int) -> str:
        """Convert playtime from minutes to hours string.

        Args:
            playtime_minutes: Playtime in minutes

        Returns:
            Playtime as hours string (e.g., "12.5")
        """
        return str(round(playtime_minutes / 60, 2)).replace(",", ".")

    @staticmethod
    def update_game_with_playtime(
        sheet: Worksheet,
        row_number: int,
        game: SteamGame,
        date_formatter,
    ) -> None:
        """Update Excel row with game playtime and last launch date.

        Updates MY_TIME_BEAT, LAST_LAUNCH_DATE, and STATUS columns
        based on Steam game data.

        Args:
            sheet: Excel worksheet to update
            row_number: Row number (1-based) to update
            game: SteamGame instance with playtime data
            date_formatter: Function to convert epoch timestamp to date string
        """
        # Update playtime
        playtime_hours = SteamExcelFormatter.format_playtime_hours(
            game.playtime_forever
        )
        sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value = (
            playtime_hours
        )

        # Update last launch date if available
        if game.rtime_last_played is not None:
            formatted_date = date_formatter(game.rtime_last_played)
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value = (
                formatted_date
            )

        # Update status if game was not started
        current_status = sheet.cell(row=row_number, column=ExcelColumn.STATUS).value
        if current_status == "Not Started":
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Dropped"

    @staticmethod
    def reset_game_without_playtime(
        sheet: Worksheet,
        row_number: int,
    ) -> None:
        """Reset game row when playtime is zero and additional_time is 'none'.

        Sets MY_TIME_BEAT to EXCEL_NONE_VALUE, LAST_LAUNCH_DATE to
        EXCEL_DATE_NOT_SET, and STATUS to "Not Started".

        Args:
            sheet: Excel worksheet to update
            row_number: Row number (1-based) to update
        """
        additional_time = sheet.cell(
            row=row_number, column=ExcelColumn.ADDITIONAL_TIME
        ).value

        if additional_time == EXCEL_NONE_VALUE:
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value = (
                EXCEL_NONE_VALUE
            )
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value = (
                EXCEL_DATE_NOT_SET
            )
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = "Not Started"

    @staticmethod
    def update_game_row(
        sheet: Worksheet,
        row_number: int,
        game: SteamGame,
        date_formatter,
    ) -> None:
        """Update Excel row with Steam game data.

        This is the main entry point for updating a game row. It handles
        both games with playtime and games without playtime.

        Args:
            sheet: Excel worksheet to update
            row_number: Row number (1-based) to update
            game: SteamGame instance with game data
            date_formatter: Function to convert epoch timestamp to date string
        """
        if game.playtime_forever != 0:
            SteamExcelFormatter.update_game_with_playtime(
                sheet, row_number, game, date_formatter
            )
        else:
            SteamExcelFormatter.reset_game_without_playtime(sheet, row_number)
