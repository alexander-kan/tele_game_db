"""Excel formatting for HowLongToBeat synchronization.

This module provides functionality to update Excel cells with
HowLongToBeat game data, specifically average time to beat.
"""

from __future__ import annotations

import logging

from openpyxl.worksheet.worksheet import Worksheet

from ..constants import ExcelColumn

logger = logging.getLogger("game_db.excel.hltb")


class HowLongToBeatExcelFormatter:
    """Format Excel cells with HowLongToBeat game data."""

    @staticmethod
    def format_time(time_hours: float | None) -> str | None:
        """Format time value for Excel.

        Args:
            time_hours: Time in hours from HowLongToBeat

        Returns:
            Formatted time string (hours as float) or None
        """
        if time_hours is None:
            return None

        try:
            time_float = float(time_hours)
            if time_float <= 0:
                return None
            return str(time_float)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def update_game_row(
        sheet: Worksheet,
        row_number: int,
        hltb_data: dict | None,
        partial_mode: bool = False,
    ) -> None:
        """Update Excel row with HowLongToBeat game data.

        Updates AVERAGE_TIME_BEAT column.

        Args:
            sheet: Excel worksheet to update
            row_number: Row number (1-based) to update
            hltb_data: Dictionary with HowLongToBeat data, or None if game not found
            partial_mode: If True, write "0" when game not found.
                         If False, only write "0" if field is empty
                         (preserve existing value).
        """
        # If game not found (hltb_data is None)
        if hltb_data is None:
            if partial_mode:
                # Partial mode: always write "0" when game not found
                sheet.cell(
                    row=row_number, column=ExcelColumn.AVERAGE_TIME_BEAT
                ).value = "0"
                logger.info(
                    "[HLTB_FORMATTER] Row %d: Game not found, "
                    "wrote '0' (partial_mode=True)",
                    row_number,
                )
            else:
                # Full mode: only write "0" if field is currently empty
                current_value = sheet.cell(
                    row=row_number, column=ExcelColumn.AVERAGE_TIME_BEAT
                ).value
                current_value_str = (
                    str(current_value).strip() if current_value else ""
                )
                if not current_value_str:
                    # Field is empty, write "0"
                    sheet.cell(
                        row=row_number, column=ExcelColumn.AVERAGE_TIME_BEAT
                    ).value = "0"
                    logger.info(
                        "[HLTB_FORMATTER] Row %d: Game not found, "
                        "wrote '0' (field was empty)",
                        row_number,
                    )
                else:
                    # Field has value, preserve it
                    logger.info(
                        "[HLTB_FORMATTER] Row %d: Game not found, "
                        "preserved existing value: %s",
                        row_number,
                        current_value_str,
                    )
            return

        # Game found: extract time data
        # Use main_story if available, otherwise use completionist
        time_hours = hltb_data.get("main_story") or hltb_data.get(
            "completionist"
        ) or hltb_data.get("main_extra") or hltb_data.get("all_styles")

        formatted_time = HowLongToBeatExcelFormatter.format_time(time_hours)
        if formatted_time is not None:
            sheet.cell(
                row=row_number, column=ExcelColumn.AVERAGE_TIME_BEAT
            ).value = formatted_time
            logger.info(
                "[HLTB_FORMATTER] Row %d: Updated AVERAGE_TIME_BEAT: %s hours",
                row_number,
                formatted_time,
            )
        else:
            logger.debug(
                "[HLTB_FORMATTER] Row %d: No valid time data to update",
                row_number,
            )
