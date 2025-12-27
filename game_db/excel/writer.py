"""Excel file writing functionality."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..constants import ExcelColumn, ExcelRowIndex
from .models import GameRow

if TYPE_CHECKING:
    from pathlib import Path


class ExcelWriter:
    """Write game data to Excel files."""

    @staticmethod
    def write_game_row(
        sheet: Worksheet,
        row_number: int,
        game_row: GameRow | list,
        xlsx_path: str | Path,
        workbook: Workbook,
    ) -> None:
        """Write a game row to Excel sheet.

        Args:
            sheet: OpenPyXL Worksheet object
            row_number: Row number to write to (1-based)
            game_row: GameRow instance or list of values
            xlsx_path: Path to Excel file (for saving)
            workbook: OpenPyXL Workbook object (for saving)
        """
        if isinstance(game_row, GameRow):
            row_data = game_row.to_list()
        else:
            row_data = game_row

        # Write each column value
        if len(row_data) > ExcelRowIndex.GAME_NAME:
            sheet.cell(row=row_number, column=ExcelColumn.GAME_NAME).value = row_data[
                ExcelRowIndex.GAME_NAME
            ]
        if len(row_data) > ExcelRowIndex.PLATFORMS:
            sheet.cell(row=row_number, column=ExcelColumn.PLATFORMS).value = row_data[
                ExcelRowIndex.PLATFORMS
            ]
        if len(row_data) > ExcelRowIndex.STATUS:
            sheet.cell(row=row_number, column=ExcelColumn.STATUS).value = row_data[
                ExcelRowIndex.STATUS
            ]
        if len(row_data) > ExcelRowIndex.RELEASE_DATE:
            sheet.cell(row=row_number, column=ExcelColumn.RELEASE_DATE).value = (
                row_data[ExcelRowIndex.RELEASE_DATE]
            )
        if len(row_data) > ExcelRowIndex.PRESS_SCORE:
            sheet.cell(row=row_number, column=ExcelColumn.PRESS_SCORE).value = row_data[
                ExcelRowIndex.PRESS_SCORE
            ]
        if len(row_data) > ExcelRowIndex.USER_SCORE:
            sheet.cell(row=row_number, column=ExcelColumn.USER_SCORE).value = row_data[
                ExcelRowIndex.USER_SCORE
            ]
        if len(row_data) > ExcelRowIndex.MY_SCORE:
            sheet.cell(row=row_number, column=ExcelColumn.MY_SCORE).value = row_data[
                ExcelRowIndex.MY_SCORE
            ]
        if len(row_data) > ExcelRowIndex.METACRITIC_URL:
            sheet.cell(row=row_number, column=ExcelColumn.METACRITIC_URL).value = (
                row_data[ExcelRowIndex.METACRITIC_URL]
            )
        if len(row_data) > ExcelRowIndex.AVERAGE_TIME_BEAT:
            sheet.cell(row=row_number, column=ExcelColumn.AVERAGE_TIME_BEAT).value = (
                row_data[ExcelRowIndex.AVERAGE_TIME_BEAT]
            )
        if len(row_data) > ExcelRowIndex.TRAILER_URL:
            sheet.cell(row=row_number, column=ExcelColumn.TRAILER_URL).value = row_data[
                ExcelRowIndex.TRAILER_URL
            ]
        if len(row_data) > ExcelRowIndex.MY_TIME_BEAT:
            sheet.cell(row=row_number, column=ExcelColumn.MY_TIME_BEAT).value = (
                row_data[ExcelRowIndex.MY_TIME_BEAT]
            )
        if len(row_data) > ExcelRowIndex.LAST_LAUNCH_DATE:
            sheet.cell(row=row_number, column=ExcelColumn.LAST_LAUNCH_DATE).value = (
                row_data[ExcelRowIndex.LAST_LAUNCH_DATE]
            )
        if len(row_data) > ExcelRowIndex.ADDITIONAL_TIME:
            sheet.cell(row=row_number, column=ExcelColumn.ADDITIONAL_TIME).value = (
                row_data[ExcelRowIndex.ADDITIONAL_TIME]
            )

    @staticmethod
    def update_init_games_sheet(
        workbook: Workbook,
        game_row: GameRow | list,
        row_number: int,
        xlsx_path: str | Path,
    ) -> None:
        """Update a row in init_games sheet.

        Args:
            workbook: OpenPyXL Workbook object
            game_row: GameRow instance or list of values
            xlsx_path: Path to Excel file
            row_number: Row number in init_games sheet
        """
        init_sheet = workbook["init_games"]
        ExcelWriter.write_game_row(
            init_sheet, row_number, game_row, xlsx_path, workbook
        )
        workbook.save(str(xlsx_path))

    @staticmethod
    def append_to_init_games_sheet(
        workbook: Workbook,
        game_row: GameRow | list,
        xlsx_path: str | Path,
    ) -> None:
        """Append a new row to init_games sheet.

        Args:
            workbook: OpenPyXL Workbook object
            game_row: GameRow instance or list of values
            xlsx_path: Path to Excel file
        """
        init_sheet = workbook["init_games"]
        new_row_number = init_sheet.max_row + 1
        ExcelWriter.write_game_row(
            init_sheet, new_row_number, game_row, xlsx_path, workbook
        )
        workbook.save(str(xlsx_path))

    @staticmethod
    def save_workbook(workbook: Workbook, file_path: str | Path) -> None:
        """Save workbook to file.

        Args:
            workbook: OpenPyXL Workbook object
            file_path: Path to save the workbook
        """
        workbook.save(str(file_path))
