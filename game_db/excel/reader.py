"""Excel file reading functionality."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..constants import ExcelColumn
from .models import GameRow

if TYPE_CHECKING:
    from pathlib import Path


class ExcelReader:
    """Read game data from Excel files."""

    @staticmethod
    def load_workbook(file_path: str | Path) -> Workbook:
        """Load Excel workbook from file.

        Args:
            file_path: Path to Excel file

        Returns:
            OpenPyXL Workbook object
        """
        return load_workbook(filename=str(file_path))

    @staticmethod
    def get_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
        """Get worksheet by name.

        Args:
            workbook: OpenPyXL Workbook object
            sheet_name: Name of the sheet to retrieve

        Returns:
            OpenPyXL Worksheet object

        Raises:
            KeyError: If sheet doesn't exist
        """
        return workbook[sheet_name]

    @staticmethod
    def read_game_rows(sheet: Worksheet, max_row: int | None = None) -> list[GameRow]:
        """Read game data rows from Excel sheet.

        Skips the first row (header) and reads data starting from row 2.

        Args:
            sheet: OpenPyXL Worksheet object
            max_row: Maximum row number to read (None = all rows)

        Returns:
            List of GameRow objects
        """
        if max_row is None:
            max_row = sheet.max_row

        game_rows: list[GameRow] = []
        # Start from row 2 to skip header (row 1)
        for i in range(2, max_row + 1):
            row_data: list = []
            for col in range(
                ExcelColumn.GAME_NAME, ExcelColumn.ADDITIONAL_TIME + 1
            ):
                cell_value = sheet.cell(row=i, column=col).value
                row_data.append(cell_value)
            # Only add non-empty rows
            if any(cell is not None for cell in row_data):
                game_row = GameRow.from_list(row_data)
                game_rows.append(game_row)
        return game_rows

    @staticmethod
    def find_row_by_game_name(
        sheet: Worksheet, game_name: str
    ) -> int | None:
        """Find row index in sheet by game name.

        Args:
            sheet: OpenPyXL Worksheet object
            game_name: Name of the game to find

        Returns:
            Row number (1-based) if found, None otherwise
        """
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(
                row=row, column=ExcelColumn.GAME_NAME
            ).value
            if cell_value == game_name:
                return row
        return None
