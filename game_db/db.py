"""Database management and high-level DB operations facade."""

from __future__ import annotations

import logging
import sqlite3
from pathlib import Path
from sqlite3 import Error
from time import localtime, strftime

from openpyxl import Workbook, load_workbook

from . import steam_api
from .config import (
    SettingsConfig,
    TokensConfig,
    load_column_table_names_config,
    load_settings_config,
    load_similarity_thresholds_config,
    load_table_names_config,
    load_tokens_config,
    load_values_dictionaries_config,
)
from .constants import ExcelColumn
from .db_dictionaries import DictionariesBuilder
from .db_excel import ExcelImporter
from .excel.hltb_formatter import HowLongToBeatExcelFormatter
from .excel.metacritic_formatter import MetacriticExcelFormatter
from .excel.steam_formatter import SteamExcelFormatter
from .hltb_client import HowLongToBeatClient
from .metacritic_search import search_metacritic_game_url
from .MetaCriticScraper import MetaCriticScraper
from .similarity_search import SimilarityMatch, find_closest_match
from .types import SteamGame

logger = logging.getLogger("game_db.sql")
_settings_cfg = load_settings_config()


class DatabaseManager:
    """Low-level work with SQLite and SQL scripts."""

    @staticmethod
    def create_connection(db_file: str | Path) -> sqlite3.Connection | None:
        """Create a database connection to a SQLite database."""
        conn = None
        try:
            conn = sqlite3.connect(str(db_file))
        except Error:
            logger.exception("Failed to connect to SQLite DB '%s'", db_file)
        return conn

    @staticmethod
    def create_table(conn: sqlite3.Connection, create_table_sql: str) -> None:
        """Create a table from the create_table_sql statement."""
        try:
            cursor = conn.cursor()
            cursor.execute(create_table_sql)
        except Error:
            logger.exception("Failed to create table with SQL: %s", create_table_sql)

    def execute_scripts_from_sql_file(
        self, sql_file: str | Path, db_file: str | Path
    ) -> None:
        """Execute SQL statements from a file using executescript.

        This method uses SQLite's executescript() which properly handles
        multiple SQL statements separated by semicolons, including those
        within string literals or comments.

        Args:
            sql_file: Path to SQL file to execute
            db_file: Path to SQLite database file
        """
        conn = self.create_connection(db_file)
        if conn is None:
            logger.error(
                "Could not obtain DB connection for '%s', " "skipping SQL file '%s'",
                db_file,
                sql_file,
            )
            return

        try:
            with open(sql_file, "r", encoding="utf-8") as fd:
                sql_file_content = fd.read()

            # Count approximate number of statements (for logging)
            # This is approximate as executescript handles it properly
            statement_count = sql_file_content.count(";")
            logger.info(
                "[SQL_EXECUTION] Executing SQL file '%s' on database '%s' "
                "(approximately %d statements)",
                sql_file,
                db_file,
                statement_count,
            )

            conn.executescript(sql_file_content)
            conn.commit()

            logger.info(
                "[SQL_EXECUTION] Successfully executed SQL file '%s' "
                "(approximately %d statements)",
                sql_file,
                statement_count,
            )
        except Error:
            logger.error(
                "[SQL_EXECUTION] Failed to execute SQL file '%s' on database '%s'",
                sql_file,
                db_file,
                exc_info=True,
            )
        finally:
            conn.close()


class SteamSynchronizer:
    """Synchronize Excel with Steam API and trigger DB recreation."""

    def __init__(
        self,
        tokens: TokensConfig,
        excel_importer: ExcelImporter,
        db_manager: DatabaseManager,
        settings: SettingsConfig,
        steam_client=None,  # Optional SteamAPI for dependency injection
    ) -> None:
        self.tokens = tokens
        self.excel_importer = excel_importer
        self.db_manager = db_manager
        self.settings = settings
        self.steam_client = steam_client or steam_api.SteamAPI()

    def epoch_date_convert(self, date: int) -> str:
        """Convert epoch seconds to 'Month DD, YYYY'."""
        date_db = strftime("%Y-%m-%d", localtime(date))
        date_list: list[str] = []
        for part in date_db.split("-"):
            date_list.append(part)
        # Simple month mapping - can be extracted to a helper if needed
        month_map = {
            "01": "January",
            "02": "February",
            "03": "March",
            "04": "April",
            "05": "May",
            "06": "June",
            "07": "July",
            "08": "August",
            "09": "September",
            "10": "October",
            "11": "November",
            "12": "December",
        }
        month_name = month_map.get(date_list[1], date_list[1])
        return f"{month_name} " f"{date_list[2]}, {date_list[0]}"

    def _load_workbook(self, xlsx_path: str | Path) -> Workbook:
        """Load Excel workbook from file.

        Args:
            xlsx_path: Path to Excel file

        Returns:
            OpenPyXL Workbook instance
        """
        logger.debug("[STEAM_SYNC] Loading Excel workbook from: %s", xlsx_path)
        return load_workbook(filename=xlsx_path)

    def _match_steam_games_with_excel(
        self, workbook: Workbook, steam_games: list[SteamGame]
    ) -> tuple[list[tuple[SteamGame, int]], list[str]]:
        """Match Steam games with Excel rows.

        Args:
            workbook: OpenPyXL Workbook instance
            steam_games: List of SteamGame instances

        Returns:
            Tuple of (matched_games, missing_games) where:
            - matched_games: List of (SteamGame, row_number) tuples
            - missing_games: List of game names not found in Excel
        """
        matched: list[tuple[SteamGame, int]] = []
        missing: list[str] = []

        for game in steam_games:
            sheet = workbook["init_games"]
            row_number = self.excel_importer.reader.find_row_by_game_name(
                sheet, game.name
            )
            if row_number:
                matched.append((game, row_number))
            else:
                missing.append(game.name)

        logger.debug(
            "Matched %d games, %d missing from Excel",
            len(matched),
            len(missing),
        )
        return matched, missing

    def _update_excel_with_steam_data(
        self, workbook: Workbook, matched_games: list[tuple[SteamGame, int]]
    ) -> None:
        """Update Excel cells with Steam game data.

        Args:
            workbook: OpenPyXL Workbook instance
            matched_games: List of (SteamGame, row_number) tuples
        """
        sheet = workbook["init_games"]
        for game, row_number in matched_games:
            SteamExcelFormatter.update_game_row(
                sheet, row_number, game, self.epoch_date_convert
            )
        logger.debug("[STEAM_SYNC] Updated %d game rows in Excel", len(matched_games))

    def _save_workbook(self, workbook: Workbook, xlsx_path: str | Path) -> None:
        """Save workbook to file.

        Args:
            workbook: OpenPyXL Workbook instance
            xlsx_path: Path to save the workbook
        """
        logger.debug("[STEAM_SYNC] Saving workbook to: %s", xlsx_path)
        workbook.save(str(xlsx_path))

    def _get_all_game_names_from_excel(self, workbook: Workbook) -> list[str]:
        """Get all game names from Excel for similarity search.

        Args:
            workbook: OpenPyXL Workbook instance

        Returns:
            List of all game names in Excel
        """
        sheet = workbook["init_games"]
        game_names: list[str] = []

        for row in range(2, sheet.max_row + 1):
            game_name_cell = sheet.cell(row=row, column=ExcelColumn.GAME_NAME)
            if game_name_cell.value:
                game_name = str(game_name_cell.value).strip()
                if game_name:
                    game_names.append(game_name)

        return game_names

    def _find_similar_games(
        self, missing_games: list[str], workbook: Workbook
    ) -> list[SimilarityMatch]:
        """Find similar games in database for missing Steam games.

        Args:
            missing_games: List of game names from Steam API not found in DB
            workbook: OpenPyXL Workbook instance

        Returns:
            List of SimilarityMatch objects
        """
        if not missing_games:
            return []

        # Get all game names from Excel
        all_game_names = self._get_all_game_names_from_excel(workbook)

        # Load similarity thresholds
        thresholds = load_similarity_thresholds_config()

        # Find closest matches for each missing game
        matches: list[SimilarityMatch] = []
        for missing_game in missing_games:
            match = find_closest_match(
                missing_game,
                all_game_names,
                thresholds,
                length_diff_threshold=thresholds.length_diff_threshold,
            )
            matches.append(match)

        return matches

    def _log_missing_games(self, missing_games: list[str]) -> None:
        """Log games from Steam that were not found in Excel.

        Args:
            missing_games: List of game names not found in Excel
        """
        if missing_games:
            logger.info(
                "Games from Steam not found in DB: %s", ", ".join(missing_games)
            )

    def _log_games_in_db_not_in_steam(
        self,
        workbook: Workbook,
        steam_game_list: list[steam_api.SteamGame],
    ) -> None:
        """Log games from DB with Steam platform that were not found in Steam API.

        Args:
            workbook: OpenPyXL Workbook instance
            steam_game_list: List of SteamGame instances from Steam API
        """
        sheet = workbook["init_games"]
        games_in_db_not_in_steam: list[str] = []

        # Iterate through all rows (starting from row 2, row 1 is header)
        for row in range(2, sheet.max_row + 1):
            game_name_cell = sheet.cell(row=row, column=ExcelColumn.GAME_NAME)
            platforms_cell = sheet.cell(row=row, column=ExcelColumn.PLATFORMS)

            # Skip if game name is empty
            if not game_name_cell.value:
                continue

            game_name = str(game_name_cell.value).strip()
            if not game_name:
                continue

            # Check if platforms cell contains "Steam"
            platforms_str = ""
            if platforms_cell.value:
                platforms_str = str(platforms_cell.value).strip()

            # Check if "Steam" is in platforms (case-insensitive)
            if "Steam" in platforms_str or "steam" in platforms_str.lower():
                # Check if this game is not in Steam API response
                # Use the same matching logic as find_row_by_game_name:
                # exact match (cell_value == game_name)
                # But in reverse: check if DB game_name exists in Steam API
                # Normalize both names by stripping whitespace for comparison
                found_in_steam = False
                game_name_normalized = game_name.strip()
                for steam_game in steam_game_list:
                    # Normalize Steam game name and compare
                    steam_name_normalized = steam_game.name.strip()
                    if steam_name_normalized == game_name_normalized:
                        found_in_steam = True
                        break

                if not found_in_steam:
                    games_in_db_not_in_steam.append(game_name)

        if games_in_db_not_in_steam:
            logger.info(
                "Games in DB with Steam platform not found in Steam API: %s",
                ", ".join(games_in_db_not_in_steam),
            )

    def _recreate_database(self, xlsx_path: str | Path) -> bool:
        """Recreate database by executing SQL scripts and importing games.

        Args:
            xlsx_path: Path to Excel file with game data

        Returns:
            True if operation succeeded, False otherwise
        """
        logger.info(
            "[DB_RECREATION] Starting database recreation from Excel file: %s",
            xlsx_path,
        )
        db_files = self.settings.db_files

        # Drop existing tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_drop_tables, db_files.sqlite_db_file
        )

        # Create new tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_create_tables, db_files.sqlite_db_file
        )

        # Generate dictionaries SQL file dynamically from values_dictionaries.ini
        table_names = load_table_names_config()
        column_table_names = load_column_table_names_config()
        values_dictionaries = load_values_dictionaries_config()
        dictionaries_builder = DictionariesBuilder(
            table_names, column_table_names, values_dictionaries
        )
        dictionaries_builder.create_dml_dictionaries(str(db_files.sql_dictionaries))

        # Populate dictionaries
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_dictionaries, db_files.sqlite_db_file
        )

        # Import games from Excel using ExcelImporter
        return self.excel_importer.add_games(str(xlsx_path), "full")

    def synchronize_steam_games(
        self, xlsx_path: str
    ) -> tuple[bool, list[SimilarityMatch]]:
        """Sync playtime and dates from Steam to Excel and then recreate DB.

        This method orchestrates the synchronization process:
        1. Loads Excel workbook
        2. Fetches games from Steam API
        3. Matches Steam games with Excel rows
        4. Updates Excel cells with Steam data
        5. Saves Excel file
        6. Finds similar games for missing ones
        7. Logs missing games
        8. Recreates database

        Args:
            xlsx_path: Path to Excel file to update

        Returns:
            Tuple of (success: bool, similarity_matches: list[SimilarityMatch])
        """
        # Load Excel workbook
        workbook = self._load_workbook(xlsx_path)

        # Fetch games from Steam
        steam_game_list = self.steam_client.get_all_games(self.tokens.steam_id)
        logger.info(
            "[STEAM_SYNC] Fetched %d games from Steam API (SteamID: %s)",
            len(steam_game_list),
            self.tokens.steam_id,
        )

        # Match Steam games with Excel rows
        matched_games, missing_games = self._match_steam_games_with_excel(
            workbook, steam_game_list
        )

        # Update Excel cells with Steam data
        self._update_excel_with_steam_data(workbook, matched_games)

        # Save workbook
        self._save_workbook(workbook, xlsx_path)

        # Find similar games for missing ones
        similarity_matches = self._find_similar_games(missing_games, workbook)

        # Log missing games
        self._log_missing_games(missing_games)

        # Log games in DB with Steam platform not found in Steam API
        self._log_games_in_db_not_in_steam(workbook, steam_game_list)

        # Recreate database
        success = self._recreate_database(xlsx_path)
        return success, similarity_matches

    def check_steam_games(self, xlsx_path: str) -> tuple[bool, list[SimilarityMatch]]:
        """Check which games from Steam are missing in database.

        This method only checks without updating anything:
        1. Loads Excel workbook
        2. Fetches games from Steam API
        3. Matches Steam games with Excel rows
        4. Finds similar games for missing ones
        5. Returns list of SimilarityMatch objects

        Args:
            xlsx_path: Path to Excel file

        Returns:
            Tuple of (success: bool, similarity_matches: list[SimilarityMatch])
        """
        # Load Excel workbook
        workbook = self._load_workbook(xlsx_path)

        # Fetch games from Steam
        steam_game_list = self.steam_client.get_all_games(self.tokens.steam_id)
        logger.info(
            "[STEAM_CHECK] Fetched %d games from Steam API (SteamID: %s)",
            len(steam_game_list),
            self.tokens.steam_id,
        )

        # Match Steam games with Excel rows
        matched_games, missing_games = self._match_steam_games_with_excel(
            workbook, steam_game_list
        )

        # Find similar games for missing ones
        similarity_matches = self._find_similar_games(missing_games, workbook)

        logger.info(
            "[STEAM_CHECK] Found %d missing games from Steam",
            len(missing_games),
        )

        return True, similarity_matches

    def add_steam_games_to_excel(self, xlsx_path: str, game_names: list[str]) -> bool:
        """Add Steam games to Excel with data from Steam API.

        Finds the first free row (where GAME_NAME is empty) and starts adding games from there.

        For each game:
        - Game name (from Steam)
        - Platform: "Steam"
        - Release date: EXCEL_DATE_NOT_SET (December 12, 4712)
        - MY_SCORE: "none"
        - ADDITIONAL_TIME: "none"

        If game was ever launched (has rtime_last_played in Steam API):
        - Status: "Dropped" (via SteamExcelFormatter.update_game_with_playtime)
        - MY_TIME_BEAT: from Steam API (playtime_forever converted to hours)
        - LAST_LAUNCH_DATE: from Steam API (rtime_last_played converted to date)

        If game was never launched:
        - Status: "Not Started"
        - MY_TIME_BEAT: "none"
        - LAST_LAUNCH_DATE: EXCEL_DATE_NOT_SET

        Args:
            xlsx_path: Path to Excel file to update
            game_names: List of game names to add

        Returns:
            True if operation succeeded, False otherwise
        """
        from .constants import EXCEL_DATE_NOT_SET, EXCEL_NONE_VALUE

        # Fetch Steam games to get rtime_last_played for each game
        steam_game_list = self.steam_client.get_all_games(self.tokens.steam_id)
        logger.info(
            "[STEAM_ADD] Fetched %d games from Steam API to get last played dates",
            len(steam_game_list),
        )

        # Create a mapping of game name -> SteamGame for quick lookup
        steam_game_map: dict[str, SteamGame] = {}
        for game in steam_game_list:
            steam_game_map[game.name] = game

        workbook = self._load_workbook(xlsx_path)
        sheet = workbook["init_games"]

        # Find the first free row (first row where GAME_NAME is empty)
        # Start from row 2 (row 1 is header)
        first_free_row = 2
        while first_free_row <= sheet.max_row:
            game_name_cell = sheet.cell(
                row=first_free_row, column=ExcelColumn.GAME_NAME
            ).value
            if not game_name_cell or str(game_name_cell).strip() == "":
                break
            first_free_row += 1
        # If no free row found, use next row after max_row
        if first_free_row > sheet.max_row:
            first_free_row = sheet.max_row + 1

        next_row = first_free_row
        logger.info(
            "[STEAM_ADD] Starting to add games from row %d (first free row)",
            next_row,
        )

        for game_name in game_names:
            # Add game name
            sheet.cell(row=next_row, column=ExcelColumn.GAME_NAME).value = game_name
            # Add platform
            sheet.cell(row=next_row, column=ExcelColumn.PLATFORMS).value = "Steam"
            # Add status
            sheet.cell(row=next_row, column=ExcelColumn.STATUS).value = "Not Started"
            # Add default date for release_date
            sheet.cell(row=next_row, column=ExcelColumn.RELEASE_DATE).value = (
                EXCEL_DATE_NOT_SET
            )

            # Add MY_SCORE = "none"
            sheet.cell(row=next_row, column=ExcelColumn.MY_SCORE).value = (
                EXCEL_NONE_VALUE
            )

            # Get Steam game data
            steam_game = steam_game_map.get(game_name)

            # Check if game was ever launched (has rtime_last_played)
            if steam_game and steam_game.rtime_last_played:
                # Game was launched - apply full update logic
                # Use SteamExcelFormatter to update playtime, last launch date, and status
                SteamExcelFormatter.update_game_with_playtime(
                    sheet, next_row, steam_game, self.epoch_date_convert
                )
                logger.info(
                    "[STEAM_ADD] Game %s was launched, applied full update logic "
                    "(playtime: %d min, last played: %d)",
                    game_name,
                    steam_game.playtime_forever,
                    steam_game.rtime_last_played,
                )
            else:
                # Game was never launched - set default values
                # Add MY_TIME_BEAT = "none"
                sheet.cell(row=next_row, column=ExcelColumn.MY_TIME_BEAT).value = (
                    EXCEL_NONE_VALUE
                )
                # Add LAST_LAUNCH_DATE = default
                sheet.cell(row=next_row, column=ExcelColumn.LAST_LAUNCH_DATE).value = (
                    EXCEL_DATE_NOT_SET
                )
                logger.debug(
                    "[STEAM_ADD] Game %s was never launched, using defaults",
                    game_name,
                )

            # Add ADDITIONAL_TIME = "none" (always)
            sheet.cell(row=next_row, column=ExcelColumn.ADDITIONAL_TIME).value = (
                EXCEL_NONE_VALUE
            )

            next_row += 1

        logger.info(
            "[STEAM_ADD] Added %d games to Excel starting from row %d",
            len(game_names),
            first_free_row,
        )

        # Save workbook
        self._save_workbook(workbook, xlsx_path)

        # Recreate database
        return self._recreate_database(xlsx_path)


class MetacriticSynchronizer:
    """Synchronize Excel with Metacritic and trigger DB recreation."""

    def __init__(
        self,
        excel_importer: ExcelImporter,
        db_manager: DatabaseManager,
        settings: SettingsConfig,
        test_mode: bool = False,
    ) -> None:
        """Initialize Metacritic synchronizer.

        Args:
            excel_importer: ExcelImporter instance for reading Excel files
            db_manager: DatabaseManager instance for DB operations
            settings: SettingsConfig instance
            test_mode: If True, limit to 20 games for testing
        """
        self.excel_importer = excel_importer
        self.db_manager = db_manager
        self.settings = settings
        self.test_mode = test_mode

    def _load_workbook(self, xlsx_path: str | Path) -> Workbook:
        """Load Excel workbook from file.

        Args:
            xlsx_path: Path to Excel file

        Returns:
            OpenPyXL Workbook instance
        """
        logger.debug("[METACRITIC_SYNC] Loading Excel workbook from: %s", xlsx_path)
        return load_workbook(filename=xlsx_path)

    def _get_games_for_sync(
        self, workbook: Workbook, partial_mode: bool = False
    ) -> list[tuple[str, int, bool]]:
        """Get games from Excel for Metacritic synchronization.

        Returns games with existing URLs and games without URLs (to search).

        Args:
            workbook: OpenPyXL Workbook instance
            partial_mode: If True, only include games missing press_score or user_score

        Returns:
            List of (metacritic_url_or_name, row_number, has_url) tuples
            where has_url is True if URL exists, False if needs search
        """
        sheet = workbook["init_games"]
        games: list[tuple[str, int, bool]] = []

        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            game_name = sheet.cell(row=row, column=ExcelColumn.GAME_NAME).value
            metacritic_url = sheet.cell(
                row=row, column=ExcelColumn.METACRITIC_URL
            ).value
            press_score = sheet.cell(row=row, column=ExcelColumn.PRESS_SCORE).value
            user_score = sheet.cell(row=row, column=ExcelColumn.USER_SCORE).value

            # Skip empty rows
            if not game_name:
                continue

            # In partial mode, skip games that already have both scores
            if partial_mode:
                press_score_str = str(press_score).strip() if press_score else ""
                user_score_str = str(user_score).strip() if user_score else ""
                # Skip if both scores are filled
                if press_score_str and user_score_str:
                    continue

            game_name_str = str(game_name).strip()

            # Check if game has Metacritic URL
            if metacritic_url and str(metacritic_url).strip():
                url_str = str(metacritic_url).strip()
                if url_str and "metacritic.com" in url_str.lower():
                    games.append((url_str, row, True))
            else:
                # Game without URL - will search by name
                games.append((game_name_str, row, False))

        logger.debug(
            "[METACRITIC_SYNC] Found %d games for synchronization "
            "(%d with URL, %d without URL, partial_mode: %s)",
            len(games),
            sum(1 for _, _, has_url in games if has_url),
            sum(1 for _, _, has_url in games if not has_url),
            partial_mode,
        )
        return games

    def _update_excel_with_metacritic_data(
        self,
        workbook: Workbook,
        games_data: list[tuple[dict, int]],
    ) -> None:
        """Update Excel cells with Metacritic game data.

        Args:
            workbook: OpenPyXL Workbook instance
            games_data: List of (metacritic_data, row_number) tuples
        """
        sheet = workbook["init_games"]
        for metacritic_data, row_number in games_data:
            MetacriticExcelFormatter.update_game_row(sheet, row_number, metacritic_data)
        logger.debug(
            "[METACRITIC_SYNC] Updated %d game rows in Excel",
            len(games_data),
        )

    def _save_workbook(self, workbook: Workbook, xlsx_path: str | Path) -> None:
        """Save workbook to file.

        Args:
            workbook: OpenPyXL Workbook instance
            xlsx_path: Path to save the workbook
        """
        logger.debug("[METACRITIC_SYNC] Saving workbook to: %s", xlsx_path)
        workbook.save(str(xlsx_path))

    def _recreate_database(self, xlsx_path: str | Path) -> bool:
        """Recreate database by executing SQL scripts and importing games.

        Args:
            xlsx_path: Path to Excel file with game data

        Returns:
            True if operation succeeded, False otherwise
        """
        logger.info(
            "[DB_RECREATION] Starting database recreation from Excel file: %s",
            xlsx_path,
        )
        db_files = self.settings.db_files

        # Drop existing tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_drop_tables, db_files.sqlite_db_file
        )

        # Create new tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_create_tables, db_files.sqlite_db_file
        )

        # Generate dictionaries SQL file dynamically from values_dictionaries.ini
        table_names = load_table_names_config()
        column_table_names = load_column_table_names_config()
        values_dictionaries = load_values_dictionaries_config()
        dictionaries_builder = DictionariesBuilder(
            table_names, column_table_names, values_dictionaries
        )
        dictionaries_builder.create_dml_dictionaries(str(db_files.sql_dictionaries))

        # Populate dictionaries
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_dictionaries, db_files.sqlite_db_file
        )

        # Import games from Excel using ExcelImporter
        return self.excel_importer.add_games(str(xlsx_path), "full")

    def synchronize_metacritic_games(
        self, xlsx_path: str, partial_mode: bool = False
    ) -> bool | None:
        """Sync game data from Metacritic to Excel and then recreate DB.

        This method orchestrates the synchronization process:
        1. Loads Excel workbook
        2. Gets games with Metacritic URLs
        3. Fetches data from Metacritic for each game (with 10s delay)
        4. Updates Excel cells with Metacritic data
        5. Saves Excel file
        6. Recreates database

        Args:
            xlsx_path: Path to Excel file to update
            partial_mode: If True, only sync games missing press_score or user_score

        Returns:
            True if operation succeeded, False if error, None if no data to sync
        """
        import time

        # Load Excel workbook
        workbook = self._load_workbook(xlsx_path)

        # Get games for synchronization (with and without URLs)
        games_for_sync = self._get_games_for_sync(workbook, partial_mode=partial_mode)

        # Limit to 20 games in test mode
        if self.test_mode:
            games_for_sync = games_for_sync[:20]
            logger.info(
                "[METACRITIC_SYNC] Test mode: limiting to %d games",
                len(games_for_sync),
            )

        if not games_for_sync:
            logger.info(
                "[METACRITIC_SYNC] No games found for synchronization (partial_mode: %s)",
                partial_mode,
            )
            # Return None to indicate "no data" (not an error)
            return None

        # Fetch data from Metacritic for each game
        games_data: list[tuple[dict, int]] = []
        total_games = len(games_for_sync)

        for i, (url_or_name, row_number, has_url) in enumerate(games_for_sync, 1):
            # Determine actual URL to use
            if has_url:
                # Use existing URL
                actual_url = url_or_name
                logger.info(
                    "[METACRITIC_SYNC] Processing game %d/%d (with URL): %s",
                    i,
                    total_games,
                    actual_url,
                )
            else:
                # Search for game by name
                game_name = url_or_name
                logger.info(
                    "[METACRITIC_SYNC] Processing game %d/%d (searching): %s",
                    i,
                    total_games,
                    game_name,
                )
                metacritic_url: str | None = search_metacritic_game_url(game_name)
                if not metacritic_url:
                    logger.warning(
                        "[METACRITIC_SYNC] Could not find Metacritic URL "
                        "for game: %s",
                        game_name,
                    )
                    # Wait before next request even if search failed
                    if i < total_games:
                        time.sleep(10)
                    continue

                logger.info(
                    "[METACRITIC_SYNC] Found Metacritic URL for '%s': %s",
                    game_name,
                    actual_url,
                )

            try:
                scraper = MetaCriticScraper(actual_url)
                if scraper.game:
                    # Ensure URL is set (use actual_url if scraper didn't set it)
                    if not scraper.game.get("url") or scraper.game["url"] == "":
                        scraper.game["url"] = actual_url

                    # Log what data we received
                    logger.info(
                        "[METACRITIC_SYNC] Data for row %d: release_date=%r, "
                        "critic_score=%r, user_score=%r",
                        row_number,
                        scraper.game.get("release_date"),
                        scraper.game.get("critic_score"),
                        scraper.game.get("user_score"),
                    )

                    games_data.append((scraper.game, row_number))
                    logger.debug(
                        "[METACRITIC_SYNC] Successfully fetched data for row %d",
                        row_number,
                    )
                else:
                    logger.warning(
                        "[METACRITIC_SYNC] No data fetched for URL: %s",
                        actual_url,
                    )
            except Exception as e:
                logger.exception(
                    "[METACRITIC_SYNC] Error fetching data for URL %s: %s",
                    actual_url,
                    str(e),
                )

            # Wait 10 seconds between requests (except for the last one)
            if i < total_games:
                logger.debug(
                    "[METACRITIC_SYNC] Waiting 10 seconds before next request..."
                )
                time.sleep(10)

        # Update Excel cells with Metacritic data
        if games_data:
            self._update_excel_with_metacritic_data(workbook, games_data)

        # Save workbook
        self._save_workbook(workbook, xlsx_path)

        # Recreate database
        return self._recreate_database(xlsx_path)


class HowLongToBeatSynchronizer:
    """Synchronize Excel with HowLongToBeat and trigger DB recreation."""

    def __init__(
        self,
        excel_importer: ExcelImporter,
        db_manager: DatabaseManager,
        settings: SettingsConfig,
        test_mode: bool = False,
    ) -> None:
        """Initialize HowLongToBeat synchronizer.

        Args:
            excel_importer: ExcelImporter instance for reading Excel files
            db_manager: DatabaseManager instance for DB operations
            settings: SettingsConfig instance
            test_mode: If True, limit to 20 games for testing
        """
        self.excel_importer = excel_importer
        self.db_manager = db_manager
        self.settings = settings
        self.test_mode = test_mode
        self.hltb_client = HowLongToBeatClient()

    def _load_workbook(self, xlsx_path: str | Path) -> Workbook:
        """Load Excel workbook from file.

        Args:
            xlsx_path: Path to Excel file

        Returns:
            OpenPyXL Workbook instance
        """
        logger.debug("[HLTB_SYNC] Loading Excel workbook from: %s", xlsx_path)
        return load_workbook(filename=xlsx_path)

    def _get_games_for_sync(
        self, workbook: Workbook, partial_mode: bool = False
    ) -> list[tuple[str, int]]:
        """Get games from Excel for HowLongToBeat synchronization.

        Args:
            workbook: OpenPyXL Workbook instance
            partial_mode: If True, only include games missing average_time_beat

        Returns:
            List of (game_name, row_number) tuples
        """
        sheet = workbook["init_games"]
        games: list[tuple[str, int]] = []

        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            game_name = sheet.cell(row=row, column=ExcelColumn.GAME_NAME).value
            average_time_beat = sheet.cell(
                row=row, column=ExcelColumn.AVERAGE_TIME_BEAT
            ).value

            # Skip empty rows
            if not game_name:
                continue

            # In partial mode, skip games that already have average_time_beat
            if partial_mode:
                time_str = str(average_time_beat).strip() if average_time_beat else ""
                # Skip if time is filled
                if time_str:
                    continue

            game_name_str = str(game_name).strip()
            if game_name_str:
                games.append((game_name_str, row))

        logger.debug(
            "[HLTB_SYNC] Found %d games for synchronization (partial_mode: %s)",
            len(games),
            partial_mode,
        )
        return games

    def _update_excel_with_hltb_data(
        self,
        workbook: Workbook,
        games_data: list[tuple[dict | None, int]],
        partial_mode: bool = False,
    ) -> None:
        """Update Excel cells with HowLongToBeat game data.

        Args:
            workbook: OpenPyXL Workbook instance
            games_data: List of (hltb_data or None, row_number) tuples
            partial_mode: If True, write "0" when game not found.
                         If False, only write "0" if field is empty.
        """
        sheet = workbook["init_games"]
        for hltb_data, row_number in games_data:
            HowLongToBeatExcelFormatter.update_game_row(
                sheet, row_number, hltb_data, partial_mode
            )
        logger.debug(
            "[HLTB_SYNC] Updated %d game rows in Excel (partial_mode: %s)",
            len(games_data),
            partial_mode,
        )

    def _save_workbook(self, workbook: Workbook, xlsx_path: str | Path) -> None:
        """Save workbook to file.

        Args:
            workbook: OpenPyXL Workbook instance
            xlsx_path: Path to save the workbook
        """
        logger.debug("[HLTB_SYNC] Saving workbook to: %s", xlsx_path)
        workbook.save(str(xlsx_path))

    def _recreate_database(self, xlsx_path: str | Path) -> bool:
        """Recreate database by executing SQL scripts and importing games.

        Args:
            xlsx_path: Path to Excel file with game data

        Returns:
            True if operation succeeded, False otherwise
        """
        logger.info(
            "[DB_RECREATION] Starting database recreation from Excel file: %s",
            xlsx_path,
        )
        db_files = self.settings.db_files

        # Drop existing tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_drop_tables, db_files.sqlite_db_file
        )

        # Create new tables
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_create_tables, db_files.sqlite_db_file
        )

        # Generate dictionaries SQL file dynamically from values_dictionaries.ini
        table_names = load_table_names_config()
        column_table_names = load_column_table_names_config()
        values_dictionaries = load_values_dictionaries_config()
        dictionaries_builder = DictionariesBuilder(
            table_names, column_table_names, values_dictionaries
        )
        dictionaries_builder.create_dml_dictionaries(str(db_files.sql_dictionaries))

        # Populate dictionaries
        self.db_manager.execute_scripts_from_sql_file(
            db_files.sql_dictionaries, db_files.sqlite_db_file
        )

        # Import games from Excel using ExcelImporter
        return self.excel_importer.add_games(str(xlsx_path), "full")

    def synchronize_hltb_games(
        self, xlsx_path: str, partial_mode: bool = False
    ) -> bool | None:
        """Sync game data from HowLongToBeat to Excel and then recreate DB.

        This method orchestrates the synchronization process:
        1. Loads Excel workbook
        2. Gets all games from Excel
        3. Fetches data from HowLongToBeat for each game (with 10s delay)
        4. Updates Excel cells with HowLongToBeat data
        5. Saves Excel file
        6. Recreates database

        Args:
            xlsx_path: Path to Excel file to update
            partial_mode: If True, only sync games missing average_time_beat

        Returns:
            True if operation succeeded, False if error, None if no data to sync
        """
        import time

        # Load Excel workbook
        workbook = self._load_workbook(xlsx_path)

        # Get games for synchronization
        games_for_sync = self._get_games_for_sync(workbook, partial_mode=partial_mode)

        # Limit to 20 games in test mode
        if self.test_mode:
            games_for_sync = games_for_sync[:20]
            logger.info(
                "[HLTB_SYNC] Test mode: limiting to %d games",
                len(games_for_sync),
            )

        if not games_for_sync:
            logger.info(
                "[HLTB_SYNC] No games found for synchronization (partial_mode: %s)",
                partial_mode,
            )
            # Return None to indicate "no data" (not an error)
            return None

        # Fetch data from HowLongToBeat for each game
        # Store (hltb_data or None, row_number) for all games, including not found
        games_data: list[tuple[dict | None, int]] = []
        total_games = len(games_for_sync)

        for i, (game_name, row_number) in enumerate(games_for_sync, 1):
            logger.info(
                "[HLTB_SYNC] Processing game %d/%d: %s",
                i,
                total_games,
                game_name,
            )

            hltb_data = None
            try:
                hltb_data = self.hltb_client.search_game(game_name)
                if hltb_data:
                    logger.debug(
                        "[HLTB_SYNC] Successfully fetched data for row %d",
                        row_number,
                    )
                else:
                    logger.warning(
                        "[HLTB_SYNC] No data fetched for game: %s (row %d)",
                        game_name,
                        row_number,
                    )
            except Exception as e:  # pylint: disable=broad-except
                logger.exception(
                    "[HLTB_SYNC] Error fetching data for game %s: %s",
                    game_name,
                    str(e),
                )

            # Always append, even if hltb_data is None
            # This allows formatter to handle "not found" cases
            games_data.append((hltb_data, row_number))

            # Wait 10 seconds between requests (except for the last one)
            if i < total_games:
                logger.debug("[HLTB_SYNC] Waiting 10 seconds before next request...")
                time.sleep(10)

        # Update Excel cells with HowLongToBeat data
        # Pass partial_mode to formatter so it knows how to handle None values
        if games_data:
            self._update_excel_with_hltb_data(workbook, games_data, partial_mode)

        # Save workbook
        self._save_workbook(workbook, xlsx_path)

        # Recreate database
        return self._recreate_database(xlsx_path)


class ChangeDB:
    """Facade kept for backward compatibility.

    This class is maintained for backward compatibility with existing code.
    It internally uses DatabaseService, which is the recommended way to
    perform database operations in new code.

    For new code, prefer using DatabaseService directly with typed
    configuration objects.
    """

    def __init__(self) -> None:
        """Initialize ChangeDB facade.

        Loads configuration and creates internal DatabaseService instance.
        """
        # Load configs via unified config layer
        self.settings: SettingsConfig = load_settings_config()
        self.tokens: TokensConfig = load_tokens_config()

        # Create service instance (recommended approach for new code)
        from .services.database_service import DatabaseService

        self._service = DatabaseService(self.settings, self.tokens)

        # Expose settings for backward compatibility
        self.table_names = self._service.table_names
        self.column_table_names = self._service.column_table_names
        self.values_dictionaries = self._service.values_dictionaries
        self.db_manager = self._service.db_manager
        self.dictionaries_builder = self._service.dictionaries_builder
        self.excel_importer = self._service.excel_importer
        self.steam_synchronizer = self._service.steam_synchronizer

    def recreate_db(self, xlsx: str) -> bool:
        """Drop, recreate and fill DB from Excel.

        Args:
            xlsx: Path to Excel file

        Returns:
            True if operation succeeded, False otherwise
        """
        return self._service.recreate_db(xlsx)

    def add_games(self, xlsx: str, mode: str) -> bool:
        """Add games from Excel to database.

        Args:
            xlsx: Path to Excel file
            mode: Import mode ("full" or "update")

        Returns:
            True if operation succeeded, False otherwise
        """
        return self._service.add_games(xlsx, mode)

    def create_dml_dictionaries(self, sql_dictionaries: str) -> None:
        """Generate SQL file for dictionary inserts.

        Args:
            sql_dictionaries: Path to output SQL file
        """
        self._service.create_dml_dictionaries(sql_dictionaries)

    def synchronize_steam_games(self, xlsx_path: str) -> tuple[bool, list]:
        """Synchronize Steam playtime with Excel and recreate DB.

        Args:
            xlsx_path: Path to Excel file

        Returns:
            Tuple of (success: bool, similarity_matches: list)
        """
        return self._service.synchronize_steam_games(xlsx_path)

    def check_steam_games(self, xlsx_path: str) -> tuple[bool, list]:
        """Check which games from Steam are missing in database."""
        return self._service.check_steam_games(xlsx_path)

    def add_steam_games_to_excel(self, xlsx_path: str, game_names: list[str]) -> bool:
        """Add Steam games to Excel with minimal data."""
        return self._service.add_steam_games_to_excel(xlsx_path, game_names)

    def synchronize_metacritic_games(
        self,
        xlsx_path: str,
        test_mode: bool = False,
        partial_mode: bool = False,
    ) -> bool | None:
        """Synchronize Metacritic data with Excel and recreate DB.

        Args:
            xlsx_path: Path to Excel file
            test_mode: If True, limit to 20 games for testing
            partial_mode: If True, only sync games missing press_score or user_score

        Returns:
            True if operation succeeded, False if error, None if no data to sync
        """
        return self._service.synchronize_metacritic_games(
            xlsx_path, test_mode=test_mode, partial_mode=partial_mode
        )

    def synchronize_hltb_games(
        self,
        xlsx_path: str,
        test_mode: bool = False,
        partial_mode: bool = False,
    ) -> bool | None:
        """Synchronize HowLongToBeat data with Excel and recreate DB."""
        return self._service.synchronize_hltb_games(
            xlsx_path, test_mode=test_mode, partial_mode=partial_mode
        )


def main() -> None:
    """Simple CLI entrypoint to recreate DB from default Excel.

    This function is kept for backward compatibility. For new code,
    prefer using scripts/init_db.py which provides better error handling
    and uses DatabaseService directly.

    Note: This function uses hardcoded path. Consider using
    scripts/init_db.py instead, which uses settings configuration.
    """
    from .config import load_settings_config, load_tokens_config
    from .logging_config import configure_logging
    from .services.database_service import DatabaseService

    configure_logging()
    settings = load_settings_config()
    tokens = load_tokens_config()

    # Use settings path instead of hardcoded PROJECT_ROOT
    service = DatabaseService(settings, tokens)
    service.recreate_db(settings.paths.games_excel_file)


if __name__ == "__main__":
    main()
